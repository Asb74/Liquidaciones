from __future__ import annotations

import logging
import os
import tempfile
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter

from domain.calculation_models import CalculationStatus, LiquidationResult
from domain.liquidacion_calculator import calculate_fiscal_result, calculate_taxable_base
from domain.utils import round_price
from domain.audit import audit_latest_excel_row, current_audit
from exporters.file_lock import FileLockedError, ensure_target_is_writable

logger = logging.getLogger(__name__)

SUMMARY_HEADERS = [
    "Nº Socio",
    "Socio",
    "Variedad",
    "Neto",
    "I. Bruto",
    "P. Comer.",
    "Recolec.",
    "C. Has.",
    "B/P Cal.",
    "B. Trans.",
    "B. Glob.",
    "Base Imponible",
    "P.Medio",
    "I.V.A",
    "Ret.",
    "Importe Total.",
    "Concepto Liquidación",
    "Cultivo",
    "Com. Recol. (Pts/kg)(<11,65)",
    "Com. Global (Pts/kg)(<2)",
    "Com. Trans. (Pts/kg) (<1,7)",
]

INTEGER_FORMAT = '#,##0;-#,##0;-'
MONEY_FORMAT = '#,##0.00;-#,##0.00;-'
PRICE_FORMAT = '0.00000;-0.00000;-'
PTS_KG_FORMAT = '0.00;-0.00;-'
PERCENT_FORMAT = '0"%"'

COLUMN_WIDTHS = {
    "A": 12, "B": 35, "C": 18, "D": 14, "E": 15, "F": 12, "G": 14,
    "H": 14, "I": 14, "J": 14, "K": 14, "L": 17, "M": 12, "N": 10,
    "O": 10, "P": 17, "Q": 32, "R": 16, "S": 24, "T": 23, "U": 24,
}


def _number(value, field_name: str, *, required: bool = False):
    if value is None:
        if required:
            raise ValueError(f"Falta el dato obligatorio: {field_name}")
        logger.warning("Dato no informado para %s; se deja la celda vacía", field_name)
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        raise ValueError(f"Valor numérico no válido para {field_name}: {value!r}")
    if isinstance(value, (int, float)):
        return value
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Valor numérico no válido para {field_name}: {value!r}") from exc


def _validate_result(result: LiquidationResult) -> None:
    if result is None:
        raise ValueError("No existe un cálculo actual para exportar")
    if not getattr(result, "member_results", None):
        raise ValueError("El cálculo no contiene líneas para exportar")
    for idx, member in enumerate(result.member_results, start=1):
        if member.member_id in (None, ""):
            raise ValueError(f"La línea {idx} no tiene IdSocio")
        if not member.variety:
            raise ValueError(f"La línea {idx} no tiene Variedad")
        _number(member.net_kg, f"Neto línea {idx}", required=True)
        for field_name in (
            "gross_amount", "commercial_average_price", "collection_amount", "hectare_fee_amount",
            "quality_amount", "transport_amount", "globalgap_amount", "taxable_base",
            "final_average_price", "vat_rate", "withholding_rate", "total_amount",
        ):
            _number(getattr(member, field_name), f"{field_name} línea {idx}")



def _append_economic_audit_sheet(wb: Workbook, result: LiquidationResult, summary_rows: dict[int, Decimal | None], red_fill: PatternFill) -> None:
    ws = wb.create_sheet("Auditoría económica")
    headers = [
        "Nº Socio", "Socio", "Variedad", "Importe bruto mostrado", "Importe bruto usado en fórmula",
        "Recolección", "Cuota Ha", "B/P Calidad", "B. Transporte", "B. GlobalGAP",
        "Base esperada", "Base almacenada en modelo", "Base exportada", "Diferencia",
        "Precio medio esperado", "Precio medio almacenado", "Alineado", "Advertencias",
    ]
    ws.append(headers)
    tolerance = Decimal("0.01")
    for member in result.member_results:
        exported_base = summary_rows.get(member.member_id)
        if None in (member.collection_amount, member.hectare_fee_amount, member.quality_amount, member.transport_amount, member.globalgap_amount):
            expected = None
            diff = None
            expected_price = None
            aligned = False
            warnings = "Base pendiente o en error; falta algún concepto obligatorio."
        else:
            expected = calculate_taxable_base(member.gross_amount, member.collection_amount, member.hectare_fee_amount, member.quality_amount, member.transport_amount, member.globalgap_amount)
            diff = None if member.taxable_base is None else member.taxable_base - expected
            expected_price = round_price(expected / member.net_kg) if member.net_kg else None
            aligned = (
                member.taxable_base is not None
                and exported_base is not None
                and -tolerance <= expected - member.taxable_base <= tolerance
                and -tolerance <= expected - exported_base <= tolerance
            )
            warnings = "" if aligned else "Base imponible no alineada."
        ws.append([
            member.member_id, member.member_name, member.variety, _number(member.gross_amount, "Importe bruto mostrado"), _number(member.gross_amount, "Importe bruto usado en fórmula"),
            _number(member.collection_amount, "Recolección"), _number(member.hectare_fee_amount, "Cuota Ha"), _number(member.quality_amount, "B/P Calidad"), _number(member.transport_amount, "B. Transporte"), _number(member.globalgap_amount, "B. GlobalGAP"),
            _number(expected, "Base esperada"), _number(member.taxable_base, "Base almacenada"), _number(exported_base, "Base exportada"), _number(diff, "Diferencia"),
            _number(expected_price, "Precio medio esperado"), _number(member.final_average_price, "Precio medio almacenado"), "SÍ" if aligned else "NO", warnings,
        ])
        if not aligned:
            for cell in ws[ws.max_row]:
                cell.fill = red_fill



def _append_fiscal_audit_sheet(wb: Workbook, result: LiquidationResult, summary_values: dict[int, tuple[Decimal | None, Decimal | None]], red_fill: PatternFill) -> None:
    ws = wb.create_sheet("Auditoría fiscal")
    headers = ["Nº Socio", "Socio", "Variedad", "Base imponible", "IVA %", "Factor IVA", "Importe IVA", "Importe después IVA", "Retención %", "Factor retención", "Importe retención", "Importe total esperado", "Importe total modelo", "Importe total Excel", "Neto efectivo", "P.Medio esperado", "P.Medio modelo", "P.Medio Excel", "Diferencia total", "Diferencia P.Medio", "Alineado", "Advertencias"]
    ws.append(headers)
    money_tol = Decimal("0.01")
    price_tol = Decimal("0.00001")
    for member in result.member_results:
        excel_price, excel_total = summary_values.get(member.member_id, (None, None))
        if None in (member.taxable_base, member.vat_rate, member.withholding_rate):
            expected = None
            diff_total = None
            diff_price = None
            aligned = False
            warnings = "Cálculo fiscal pendiente."
        else:
            expected = calculate_fiscal_result(member.taxable_base, member.net_kg, member.vat_rate, member.withholding_rate)
            diff_total = None if member.total_amount is None else member.total_amount - expected.total_amount
            diff_price = None if member.final_average_price is None or expected.final_average_price is None else member.final_average_price - expected.final_average_price
            aligned = (
                member.total_amount is not None and excel_total is not None and
                abs(expected.total_amount - member.total_amount) <= money_tol and
                abs(expected.total_amount - excel_total) <= money_tol and
                member.final_average_price is not None and excel_price is not None and expected.final_average_price is not None and
                abs(expected.final_average_price - member.final_average_price) <= price_tol and
                abs(expected.final_average_price - excel_price) <= price_tol
            )
            warnings = "" if aligned else "Fiscalidad no alineada."
        ws.append([
            member.member_id, member.member_name, member.variety, _number(member.taxable_base, "Base imponible"), _number(member.vat_rate, "IVA %"), _number(member.vat_factor, "Factor IVA"), _number(member.vat_amount, "Importe IVA"), _number(member.amount_after_vat, "Importe después IVA"), _number(member.withholding_rate, "Retención %"), _number(member.withholding_factor, "Factor retención"), _number(member.withholding_amount, "Importe retención"),
            _number(getattr(expected, "total_amount", None), "Importe total esperado"), _number(member.total_amount, "Importe total modelo"), _number(excel_total, "Importe total Excel"), _number(member.net_kg, "Neto efectivo"), _number(getattr(expected, "final_average_price", None), "P.Medio esperado"), _number(member.final_average_price, "P.Medio modelo"), _number(excel_price, "P.Medio Excel"), _number(diff_total, "Diferencia total"), _number(diff_price, "Diferencia P.Medio"), "SÍ" if aligned else "NO", warnings,
        ])
        if not aligned:
            for cell in ws[ws.max_row]:
                cell.fill = red_fill

def _hectare_fee_excel_value(member):
    status = getattr(member, "hectare_fee_status", None)
    if status == CalculationStatus.ERROR:
        logger.warning(
            "Cuota Ha no exportable socio=%s status=%s warnings=%s",
            member.member_id,
            getattr(status, "value", status),
            "; ".join(member.warnings),
        )
        return None
    if status in (CalculationStatus.DISABLED, CalculationStatus.NOT_APPLICABLE):
        return Decimal("0")
    return _number(member.hectare_fee_amount, "C. Has.")


def _append_variety_audit_sheet(wb: Workbook, result: LiquidationResult) -> None:
    ws = wb.create_sheet("Auditoría variedades")
    headers = ["Campaña", "Empresa", "Cultivo", "Remesa", "Valor original", "Tipo resuelto", "Grupo", "Subgrupo", "Variedades resueltas", "Nº variedades", "Estado", "Advertencias", "Usado en consulta", "Alineado"]
    ws.append(headers)
    used = tuple(result.header.variedades or ())
    used_norm = {str(v).strip().upper() for v in used}
    for res in getattr(result, "variety_audit", ()) or ():
        resolved = tuple(getattr(res, "varieties", ()) or ())
        resolved_norm = {str(v).strip().upper() for v in resolved}
        aligned = bool(resolved_norm) and resolved_norm.issubset(used_norm)
        ws.append([
            result.header.campana, result.header.empresa, result.header.cultivo, result.header.remesa_name,
            getattr(res, "source_value", ""), "GROUP" if getattr(res, "is_group", False) else getattr(res, "status", ""),
            getattr(res, "group", "") or "", getattr(res, "subgroup", "") or "", ", ".join(resolved), len(resolved),
            getattr(res, "status", ""), "; ".join(getattr(res, "warnings", ()) or ()), ", ".join(used), "SÍ" if aligned else "NO",
        ])

def export_liquidation_summary(result: LiquidationResult, path: Path) -> Path:
    _validate_result(result)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(SUMMARY_HEADERS)

    summary_base_by_member: dict[int, Decimal | None] = {}
    summary_fiscal_by_member: dict[int, tuple[Decimal | None, Decimal | None]] = {}
    for row_number, member in enumerate(result.member_results, start=2):
        hectare_excel_value = _hectare_fee_excel_value(member)
        audit = current_audit()
        if audit:
            audit.audit_excel_row(member, hectare_excel_value)
        else:
            audit_latest_excel_row(member, hectare_excel_value)
        exported_taxable_base = _number(member.taxable_base, "Base Imponible")
        summary_base_by_member[member.member_id] = exported_taxable_base
        exported_final_average_price = _number(member.final_average_price, "P.Medio")
        exported_total_amount = _number(member.total_amount, "Importe Total.")
        summary_fiscal_by_member[member.member_id] = (exported_final_average_price, exported_total_amount)
        ws.append([
            member.member_id,
            member.member_name,
            member.variety,
            _number(member.net_kg, "Neto", required=True),
            _number(member.gross_amount, "I. Bruto"),
            _number(member.commercial_average_price, "P. Comer."),
            _number(member.collection_amount, "Recolec."),
            hectare_excel_value,
            _number(member.quality_amount, "B/P Cal."),
            _number(member.transport_amount, "B. Trans."),
            _number(member.globalgap_amount, "B. Glob."),
            exported_taxable_base,
            exported_final_average_price,
            _number(member.vat_rate, "I.V.A"),
            _number(member.withholding_rate, "Ret."),
            exported_total_amount,
            result.header.remesa_name,
            result.header.cultivo,
            f'=IFERROR(166.386*G{row_number}/D{row_number},0)',
            f'=IFERROR(166.386*K{row_number}/D{row_number},0)',
            f'=IFERROR(166.386*J{row_number}/D{row_number},0)',
        ])

    detail = wb.create_sheet("Detalle ajustes")
    detail.append(["Nº Socio", "Socio", "Variedad", "Bon/Pen tarifa", "Fuente tarifa", "Hectáreas", "Precio/ha", "Cuota total socio", "Kilos efectivos totales", "Proporción €/kg", "Cuota parcial", "Ajuste redondeo"])
    for member in result.member_results:
        detail.append([member.member_id, member.member_name, member.variety, _number(member.quality_rate, "Bon/Pen tarifa"), member.quality_source, _number(member.applicable_hectares, "Hectáreas"), _number(member.hectare_fee_price, "Precio/ha"), _number(member.hectare_fee_total_member, "Cuota total socio"), _number(member.hectare_fee_total_effective_kg, "Kilos efectivos totales"), _number(member.hectare_fee_rate_per_kg, "Proporción €/kg"), _number(member.hectare_fee_amount, "Cuota parcial"), _number(member.hectare_fee_rounding_adjustment, "Ajuste redondeo")])


    parcels = wb.create_sheet("02_Parcelas")
    parcel_headers = ["IdSocio", "Nombre socio", "Boleta DEEPP", "Cultivo DEEPP", "Campaña DEEPP", "Empresa DEEPP", "CHA original", "CHA activo", "Baja DEEPP", "Boleta DParcela", "Campaña DParcela", "Empresa DParcela", "Cultivo DParcela", "IdPM", "Pol", "Par", "Rec", "SupCul DParcela", "SupRec", "SupApor", "Baja DParcela", "Incluida", "Motivo exclusión", "Clave deduplicación"]
    parcels.append(parcel_headers)
    for member in result.member_results:
        for row in getattr(member, "hectare_fee_parcels", ()):
            parcels.append([
                row.get("IdSocio", member.member_id), member.member_name, row.get("Boleta DEEPP"), row.get("Cultivo DEEPP"),
                row.get("Campaña DEEPP"), row.get("Empresa DEEPP"), row.get("CHA original"), row.get("CHA activo"), row.get("Baja DEEPP"),
                row.get("Boleta DParcela"), row.get("Campaña DParcela"), row.get("Empresa DParcela"), row.get("Cultivo DParcela"),
                row.get("IdPM"), row.get("Pol"), row.get("Par"), row.get("Rec"), _number(row.get("SupCul DParcela"), "SupCul DParcela"),
                _number(row.get("SupRec"), "SupRec"), _number(row.get("SupApor"), "SupApor"), row.get("Baja DParcela"),
                row.get("Incluida"), row.get("Motivo exclusión"), row.get("Clave deduplicación"),
            ])

    total_row = ws.max_row + 1
    ws.cell(total_row, 2, "TOTAL")
    for column in (4, 5, 7, 8, 9, 10, 11, 12, 16):
        letter = get_column_letter(column)
        ws.cell(total_row, column, f"=SUM({letter}2:{letter}{total_row - 1})")
    ws.cell(total_row, 13, f"=IFERROR(P{total_row}/D{total_row},0)")

    _style_summary(ws, total_row)
    diagnostics = wb.create_sheet("04_CuotaHa")
    diagnostics.append(["Nº Socio", "Socio", "Variedad", "Estado", "Hectáreas", "Cuota anual", "Kilos efectivos totales", "Proporción €/kg", "Neto efectivo línea", "Cuota calculada", "Advertencias"])
    for member in result.member_results:
        diagnostics.append([
            member.member_id,
            member.member_name,
            member.variety,
            getattr(member.hectare_fee_status, "value", str(member.hectare_fee_status)),
            _number(member.applicable_hectares, "Hectáreas"),
            _number(member.hectare_fee_total_member, "Cuota anual"),
            _number(member.hectare_fee_total_effective_kg, "Kilos efectivos totales"),
            _number(member.hectare_fee_rate_per_kg, "Proporción €/kg"),
            _number(member.net_kg, "Neto efectivo línea"),
            _number(member.hectare_fee_amount, "Cuota calculada"),
            "; ".join(member.warnings),
        ])


    red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    gg_ws = wb.create_sheet("Auditoría GlobalGAP")
    gg_headers = ["Nº Socio", "Socio", "Variedad", "Campaña", "Empresa", "Cultivo remesa", "Checkbox AplGlobal", "Certificado", "Certificación incoherente", "Cultivos certificados", "Cultivos no certificados", "NivelGlobal", "Índice", "Bonificación base", "Categoría", "Base utilizada", "Neto efectivo", "Neto comercial", "Kilos base", "Importe detectado", "Importe aplicado", "Importe almacenado en modelo", "Importe exportado a Resumen", "Estado", "Advertencias", "Alineado"]
    gg_ws.append(gg_headers)
    for member in result.member_results:
        audit_data = getattr(member, "globalgap_audit", None)
        exported = _number(member.globalgap_amount, "B. Glob.")
        applied = getattr(audit_data, "applied_amount", None)
        expected = applied
        aligned = expected == member.globalgap_amount == exported
        gg_ws.append([
            member.member_id, member.member_name, member.variety, result.header.campana, result.header.empresa, result.header.cultivo,
            "SÍ" if result.header.options.get("GlobalGAP") else "NO", "SÍ" if getattr(audit_data, "certified", False) else "NO", "SÍ" if getattr(audit_data, "certification_inconsistent", False) else "NO",
            ", ".join(getattr(audit_data, "certified_crops", ())), ", ".join(getattr(audit_data, "non_certified_crops", ())),
            getattr(audit_data, "level", None), _number(getattr(audit_data, "index", None), "Índice"), _number(getattr(audit_data, "bonus_rate", None), "Bonificación base"), getattr(audit_data, "category", None), getattr(audit_data, "base_type", None),
            _number(getattr(audit_data, "effective_net_kg", None), "Neto efectivo"), _number(getattr(audit_data, "commercial_net_kg", None), "Neto comercial"), _number(getattr(audit_data, "base_kg", None), "Kilos base"),
            _number(getattr(audit_data, "detected_amount", None), "Importe detectado"), _number(applied, "Importe aplicado"), _number(member.globalgap_amount, "Importe almacenado"), exported,
            getattr(getattr(audit_data, "status", None), "value", str(getattr(audit_data, "status", ""))), "; ".join(getattr(audit_data, "warnings", ())), "SÍ" if aligned else "NO",
        ])
        if not aligned:
            for cell in gg_ws[gg_ws.max_row]:
                cell.fill = red_fill

    _append_variety_audit_sheet(wb, result)
    _append_economic_audit_sheet(wb, result, summary_base_by_member, red_fill)
    _append_fiscal_audit_sheet(wb, result, summary_fiscal_by_member, red_fill)

    audit_ws = wb.create_sheet("Auditoría cuota Ha")
    audit_headers = ["Nº Socio", "Socio", "Variedad", "Campaña", "Empresa", "Cultivo remesa", "Precio €/ha", "Cultivos superficie activos", "Cultivos entrega activos", "Hectáreas aplicables", "Cuota teórica total", "Kg efectivos campaña", "Índice €/kg", "Kg efectivos remesa", "Cuota parcial calculada", "Cuota almacenada en modelo", "Cuota exportada a Resumen", "Estado", "Advertencias", "Alineado"]
    audit_ws.append(audit_headers)
    for member in result.member_results:
        audit_data = getattr(member, "hectare_fee_audit", None)
        exported = _hectare_fee_excel_value(member)
        calculated = getattr(audit_data, "line_fee", None)
        aligned = calculated == member.hectare_fee_amount == exported
        audit_ws.append([
            member.member_id, member.member_name, member.variety, result.header.campana, result.header.empresa, result.header.cultivo,
            _number(getattr(audit_data, "price_per_hectare", None), "Precio €/ha"), ", ".join(getattr(audit_data, "surface_crops", ())), ", ".join(getattr(audit_data, "delivery_crops", ())),
            _number(getattr(audit_data, "applicable_hectares", None), "Hectáreas aplicables"), _number(getattr(audit_data, "total_theoretical_fee", None), "Cuota teórica total"),
            _number(getattr(audit_data, "total_effective_kg", None), "Kg efectivos campaña"), _number(getattr(audit_data, "rate_per_kg", None), "Índice €/kg"),
            _number(getattr(audit_data, "line_effective_kg", None), "Kg efectivos remesa"), _number(calculated, "Cuota parcial calculada"),
            _number(member.hectare_fee_amount, "Cuota almacenada en modelo"), _number(exported, "Cuota exportada a Resumen"),
            getattr(getattr(audit_data, "status", member.hectare_fee_status), "value", str(member.hectare_fee_status)), "; ".join(getattr(audit_data, "warnings", member.warnings)), "SÍ" if aligned else "NO",
        ])
        if not aligned:
            for cell in audit_ws[audit_ws.max_row]:
                cell.fill = red_fill

    parcelas_ws = wb.create_sheet("Parcelas cuota Ha")
    parcelas_headers = ["Nº Socio", "Socio", "Boleta DEEPP", "Boleta DParcela", "Campaña", "Empresa", "Cultivo", "CHA original", "CHA activo", "IdPM", "Pol", "Par", "Rec", "SupCul DEEPP", "SupCul DParcela", "SupRec", "SupApor", "Incluida", "Motivo exclusión", "Clave deduplicación"]
    parcelas_ws.append(parcelas_headers)
    for member in result.member_results:
        for row in getattr(member, "hectare_fee_parcels", ()):
            parcelas_ws.append([member.member_id, member.member_name, row.get("Boleta DEEPP"), row.get("Boleta DParcela"), row.get("Campaña DParcela") or row.get("Campaña DEEPP"), row.get("Empresa DParcela") or row.get("Empresa DEEPP"), row.get("Cultivo DParcela") or row.get("Cultivo DEEPP"), row.get("CHA original"), row.get("CHA activo"), row.get("IdPM"), row.get("Pol"), row.get("Par"), row.get("Rec"), _number(row.get("SupCul DEEPP"), "SupCul DEEPP"), _number(row.get("SupCul DParcela"), "SupCul DParcela"), _number(row.get("SupRec"), "SupRec"), _number(row.get("SupApor"), "SupApor"), row.get("Incluida"), row.get("Motivo exclusión"), row.get("Clave deduplicación")])

    perceco = wb.create_sheet("Comparación Perceco")
    perceco.append(["Nº Socio", "Socio", "Variedad", "Campo", "Nueva app", "Perceco", "Diferencia", "Diferencia %", "Observación"])
    fields = ["Hectáreas", "Cuota teórica", "Kg campaña", "Índice", "Kg remesa", "Cuota parcial", "Importe bruto", "Recolección", "Transporte", "Calidad", "GlobalGAP", "Base imponible", "IVA", "Retención", "Total"]
    for member in result.member_results:
        audit_data = getattr(member, "hectare_fee_audit", None)
        values = {"Hectáreas": getattr(audit_data, "applicable_hectares", None), "Cuota teórica": getattr(audit_data, "total_theoretical_fee", None), "Kg campaña": getattr(audit_data, "total_effective_kg", None), "Índice": getattr(audit_data, "rate_per_kg", None), "Kg remesa": getattr(audit_data, "line_effective_kg", None), "Cuota parcial": member.hectare_fee_amount, "Importe bruto": member.gross_amount, "Recolección": member.collection_amount, "Transporte": member.transport_amount, "Calidad": member.quality_amount, "GlobalGAP": member.globalgap_amount, "Base imponible": member.taxable_base, "IVA": member.vat_amount, "Retención": member.withholding_amount, "Total": member.total_amount}
        for field in fields:
            perceco.append([member.member_id, member.member_name, member.variety, field, _number(values[field], field), None, None, None, ""])

    path.parent.mkdir(parents=True, exist_ok=True)
    ensure_target_is_writable(path)
    temp_name = None
    try:
        with tempfile.NamedTemporaryFile(prefix=f".{path.stem}_", suffix=path.suffix, dir=path.parent, delete=False) as tmp:
            temp_name = tmp.name
        temp_path = Path(temp_name)
        wb.save(temp_path)
        try:
            os.replace(temp_path, path)
        except PermissionError as exc:
            raise FileLockedError(path) from exc
    finally:
        if temp_name:
            temp_path = Path(temp_name)
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except OSError:
                    logger.warning("No se pudo eliminar el temporal de Excel: %s", temp_path)
    return path


def _style_summary(ws, total_row: int) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=total_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.border = border

    number_formats = {
        "A": INTEGER_FORMAT,
        "D": INTEGER_FORMAT,
        "E": MONEY_FORMAT, "G": MONEY_FORMAT, "H": MONEY_FORMAT, "I": MONEY_FORMAT,
        "J": MONEY_FORMAT, "K": MONEY_FORMAT, "L": MONEY_FORMAT, "P": MONEY_FORMAT,
        "F": PRICE_FORMAT, "M": PRICE_FORMAT,
        "N": PERCENT_FORMAT, "O": PERCENT_FORMAT,
        "S": PTS_KG_FORMAT, "T": PTS_KG_FORMAT, "U": PTS_KG_FORMAT,
    }
    for column_letter, number_format in number_formats.items():
        for cell in ws[column_letter][1:]:
            cell.number_format = number_format

    for column_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
