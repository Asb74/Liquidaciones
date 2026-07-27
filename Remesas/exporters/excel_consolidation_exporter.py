"""Excel consolidado de remesas.

El módulo sólo presenta resultados ya calculados: ninguna regla económica vive aquí.
"""
from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from exporters.excel_exporter import MONEY_FORMAT, INTEGER_FORMAT, PERCENT_FORMAT, PRICE_FORMAT

DATE_FORMAT = "DD/MM/YYYY"


def _calculation(item):
    value = item.calculation_result
    return value.result if hasattr(value, "result") else value


def _value(obj, *names):
    for name in names:
        if hasattr(obj, name):
            return getattr(obj, name)
    return None


def _valid_sum(values: Iterable):
    """Do not turn an incomplete business total into an apparently valid zero."""
    values = list(values)
    if not values or any(value is None for value in values):
        return None
    return sum(values, Decimal("0"))


def _status(values, warnings=()):
    return "Pendiente" if any(value is None for value in values) else ("Con incidencias" if warnings else "Correcto")


def _style(ws, *, money=(), integer=(), percent=(), price=(), dates=(), total_row=None):
    border = Border(*(Side(style="thin", color="D9D9D9") for _ in range(4)))
    for cell in ws[1]:
        cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="center"); cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border; cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.column in money: cell.number_format = MONEY_FORMAT
            elif cell.column in integer: cell.number_format = INTEGER_FORMAT
            elif cell.column in percent: cell.number_format = PERCENT_FORMAT
            elif cell.column in price: cell.number_format = PRICE_FORMAT
            elif cell.column in dates: cell.number_format = DATE_FORMAT
    if total_row:
        for cell in ws[total_row]:
            cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="C6E0B4")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for cells in ws.columns:
        ws.column_dimensions[get_column_letter(cells[0].column)].width = min(45, max(12, max(len(str(c.value or "")) for c in cells) + 2))


GENERAL_HEADERS = ["Orden", "Id remesa", "Nombre remesa", "Campaña", "Empresa", "Cultivo", "Tipo de liquidación", "Condición/Categoría", "Periodo desde", "Periodo hasta", "Fecha de pago", "Nº socios", "Nº entregas", "Nº variedades", "Kilos netos", "Importe comercial", "Destrío", "Destrío mesa", "Podrido", "Importe bruto", "Recolección detectada", "Recolección aplicada", "Transporte detectado", "Transporte aplicado", "Calidad", "GlobalGAP", "Cuota Ha", "Base imponible", "IVA", "Retención", "Total liquidación", "Estado", "Nº advertencias"]
MEMBER_HEADERS = ["Orden", "Id remesa", "Nombre remesa", "Campaña", "Empresa", "Cultivo", "Socio", "Agricultor", "Variedad", "Nº entregas", "Kilos netos", "Kilos comerciales", "Kilos destrío", "Kilos podrido", "Importe comercial", "Importe destrío", "Importe destrío mesa", "Importe podrido", "Importe bruto", "Recolección detectada", "Recolección aplicada", "Transporte detectado", "Transporte aplicado", "Calidad", "GlobalGAP", "Cuota Ha", "Base imponible", "Régimen fiscal", "% IVA", "Importe IVA", "% retención", "Importe retención", "Total", "Precio medio comercial", "Precio medio final", "Estado", "Advertencias"]
GRADE_HEADERS = ["Orden", "Id remesa", "Nombre remesa", "Campaña", "Empresa", "Cultivo", "Socio", "Agricultor", "Variedad", "Código calibre", "Etiqueta calibre", "Kilos", "Precio €/kg", "Importe"]
WARNING_HEADERS = ["Orden", "Id remesa", "Nombre remesa", "Socio", "Agricultor", "Variedad", "Tipo o categoría", "Descripción", "Severidad", "Origen"]


def export_consolidated_liquidation_summary(results: Sequence, failed_results: Sequence, output_path: Path, **metadata) -> Path:
    wb = Workbook(); general = wb.active; general.title = "Resumen general"; general.append(GENERAL_HEADERS)
    numeric_rows = []
    for order, item in enumerate(results, 1):
        rem, calc = item.remittance, _calculation(item); t = calc.totals; warnings = tuple(getattr(calc, "warnings", ()) or ())
        amounts = [_value(t, n) for n in ("net_kg", "commercial_amount", "gross_amount", "detected_collection_amount", "collection_amount", "detected_transport_amount", "transport_amount", "quality_amount", "globalgap_amount", "hectare_fee_amount", "taxable_base", "vat_amount", "withholding_amount", "total_amount")]
        destruction = _valid_sum(_value(m, "destruction_amount") for m in calc.member_results)
        table_destruction = _valid_sum(_value(m, "table_destruction_amount") for m in calc.member_results)
        rotten = _valid_sum(_value(m, "rotten_amount") for m in calc.member_results)
        row = [order, rem.remittance_id, rem.name, rem.campaign, rem.company, rem.crop, rem.liquidation_type, rem.category, rem.period_from, rem.period_to, rem.payment_date, item.member_count, item.delivery_count, getattr(calc, "variety_count", None), amounts[0], amounts[1], destruction, table_destruction, rotten, *amounts[2:], _status(amounts, warnings), len(warnings)]
        general.append(row); numeric_rows.append(row)
    if failed_results:
        for offset, fail in enumerate(failed_results, len(results) + 1):
            r=fail.remittance; general.append([offset,r.remittance_id,r.name,r.campaign,r.company,r.crop,r.liquidation_type,r.category,r.period_from,r.period_to,r.payment_date,None,None,None,*([None]*17),"Error",1])
    if results or failed_results:
        total=[None]*len(GENERAL_HEADERS); total[0]="TOTAL GENERAL"
        for index in range(11,31): total[index]=_valid_sum(row[index] for row in numeric_rows) if not failed_results else None
        total[31]="Parcial" if failed_results or any(row[31] != "Correcto" for row in numeric_rows) else "Correcto"
        general.append(total)
    _style(general, money=range(16,32), integer=(1,2,12,13,14,15,33), dates=(9,10,11), total_row=general.max_row if results or failed_results else None)

    detail=wb.create_sheet("Detalle por socio"); detail.append(MEMBER_HEADERS)
    grades=wb.create_sheet("Detalle calibres"); grades.append(GRADE_HEADERS)
    warnings_ws=wb.create_sheet("Advertencias"); warnings_ws.append(WARNING_HEADERS); seen=set()
    warning_order=0
    def warning(rem, member, description, origin, severity="Advertencia"):
        nonlocal warning_order
        key=(rem.remittance_id, _value(member,"member_id") if member else None, _value(member,"variety") if member else None, str(description), origin)
        if key in seen:return
        seen.add(key); warning_order += 1
        warnings_ws.append([warning_order,rem.remittance_id,rem.name,_value(member,"member_id") if member else None,_value(member,"member_name") if member else None,_value(member,"variety") if member else None,rem.category,str(description),severity,origin])
    row_order=grade_order=0
    for item in results:
        rem, calc = item.remittance, _calculation(item)
        for text in getattr(calc,"warnings",()) or (): warning(rem,None,text,"Remesa")
        for member in calc.member_results:
            row_order += 1; member_warnings=tuple(getattr(member,"warnings",()) or ())
            values=[_value(member,n) for n in ("collection_amount","transport_amount","quality_amount","globalgap_amount","hectare_fee_amount","taxable_base","vat_amount","withholding_amount","total_amount")]
            detail.append([row_order,rem.remittance_id,rem.name,rem.campaign,rem.company,rem.crop,member.member_id,member.member_name,member.variety,_value(member,"delivery_count"),_value(member,"net_kg"),_value(member,"commercial_kg","net_commercial"),_value(member,"destruction_kg","net_waste"),_value(member,"rotten_kg","net_rotten"),_value(member,"commercial_amount"),_value(member,"destruction_amount"),_value(member,"table_destruction_amount"),_value(member,"rotten_amount"),_value(member,"gross_amount"),_value(member,"detected_collection_amount"),_value(member,"collection_amount"),_value(member,"detected_transport_amount"),_value(member,"transport_amount"),_value(member,"quality_amount"),_value(member,"globalgap_amount"),_value(member,"hectare_fee_amount"),_value(member,"taxable_base"),_value(member,"fiscal_regime_name"),_value(member,"vat_rate"),_value(member,"vat_amount"),_value(member,"withholding_rate"),_value(member,"withholding_amount"),_value(member,"total_amount"),_value(member,"commercial_average_price"),_value(member,"final_average_price"),_status(values,member_warnings),"; ".join(map(str,member_warnings))])
            for text in member_warnings: warning(rem,member,text,"Socio")
            for grade in getattr(member,"grades",()) or ():
                grade_order += 1; grades.append([grade_order,rem.remittance_id,rem.name,rem.campaign,rem.company,rem.crop,member.member_id,member.member_name,member.variety,grade.code,grade.label,grade.kilograms,grade.price,grade.amount])
    _style(detail,money=range(15,28),integer=(1,2,7,10,11,12,13,14),percent=(29,31),price=(34,35))
    _style(grades,money=(14,),integer=(1,2,7,12),price=(13,))
    for fail in failed_results: warning(fail.remittance,None,fail.error_message,"Cálculo","Error")
    _style(warnings_ws,integer=(1,2,4))
    failed=wb.create_sheet("Remesas fallidas"); failed.append(["Orden","Id remesa","Nombre remesa","Error","Detalle técnico","Estado"])
    for order,item in enumerate(failed_results,1): failed.append([order,item.remittance.remittance_id,item.remittance.name,item.error_type,item.error_message,"Error"])
    _style(failed,integer=(1,2))
    output_path=Path(output_path)
    if output_path.suffix.lower() != ".xlsx": output_path=output_path.with_suffix(".xlsx")
    output_path.parent.mkdir(parents=True,exist_ok=True); wb.save(output_path); return output_path
