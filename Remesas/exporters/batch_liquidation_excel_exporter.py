from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from exporters.excel_exporter import (
    MONEY_FORMAT as SUMMARY_MONEY_FORMAT,
    INTEGER_FORMAT as SUMMARY_INTEGER_FORMAT,
    PERCENT_FORMAT as SUMMARY_PERCENT_FORMAT,
    PRICE_FORMAT,
    PTS_KG_FORMAT,
    SUMMARY_HEADERS,
    SummaryBlockResult,
    SummaryColumn,
    apply_liquidation_summary_style,
    build_liquidation_summary_rows,
    get_liquidation_summary_columns,
)

MONEY_FORMAT = '#,##0.00;-#,##0.00;-'
INTEGER_FORMAT = '#,##0;-#,##0;-'
PERCENT_FORMAT = '0"%"'
DATE_FORMAT = 'DD/MM/YYYY'


def _n(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    return value


def _sum_members(result, attr):
    total = Decimal("0")
    seen = False
    for member in getattr(result, "member_results", ()):
        value = getattr(member, attr, None)
        if value is not None:
            total += value
            seen = True
    return total if seen else None


def _style_sheet(ws, *, money_cols=(), integer_cols=(), percent_cols=(), date_cols=()):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.column in money_cols:
                cell.number_format = MONEY_FORMAT
            elif cell.column in integer_cols:
                cell.number_format = INTEGER_FORMAT
            elif cell.column in percent_cols:
                cell.number_format = PERCENT_FORMAT
            elif cell.column in date_cols:
                cell.number_format = DATE_FORMAT
    for column_cells in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(width, 12)


def _mark_total_row(ws, row: int, color: str):
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=color)


def _decimal(value) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _format_date(value):
    if value is None:
        return ""
    return value.strftime("%d/%m/%Y") if hasattr(value, "strftime") else str(value)


def _set_merged_value(ws, row: int, start_col: int, end_col: int, value: str, *, fill: str, bold: bool = True) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col, value)
    cell.font = Font(bold=bold, color="FFFFFF" if fill == "1F4E78" else "000000")
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(wrap_text=True, vertical="center")


def append_liquidation_summary_block(ws, start_row: int, *, remittance, calculation_result) -> SummaryBlockResult:
    columns = get_liquidation_summary_columns()
    max_col = len(columns)
    title = f"REMESA {remittance.remittance_id} - {remittance.name}"
    _set_merged_value(ws, start_row, 1, max_col, title, fill="1F4E78")
    ws.row_dimensions[start_row].height = 24
    meta1 = f"Campaña: {remittance.campaign} | Empresa: {remittance.company} | Cultivo: {remittance.crop}"
    meta2 = f"Periodo: {_format_date(remittance.period_from)} - {_format_date(remittance.period_to)} | Pago: {_format_date(remittance.payment_date)}"
    meta3 = f"Categoría: {remittance.category} | Tipo liquidación: {remittance.liquidation_type}"
    for offset, value in enumerate((meta1, meta2, meta3), start=1):
        _set_merged_value(ws, start_row + offset, 1, max_col, value, fill="D9EAF7", bold=True)
    header_row = start_row + 4
    for col, header in enumerate(SUMMARY_HEADERS, start=1):
        ws.cell(header_row, col, header)

    member_results = tuple(getattr(calculation_result, "member_results", ()) or ())
    if not member_results:
        info_row = header_row + 1
        _set_merged_value(ws, info_row, 1, max_col, "Sin liquidaciones válidas.", fill="FFF2CC", bold=True)
        return SummaryBlockResult(start_row, header_row, info_row, info_row, None, info_row + 3, 0, {})

    data_start = header_row + 1
    for row_values in build_liquidation_summary_rows(calculation_result, start_row=data_start):
        ws.append(row_values)
    data_end = data_start + len(member_results) - 1
    subtotal_row = data_end + 1
    ws.cell(subtotal_row, 2, f"SUBTOTAL REMESA {remittance.remittance_id}")
    for col_idx, column in enumerate(columns, start=1):
        if column.total_formula == "sum":
            letter = get_column_letter(col_idx)
            ws.cell(subtotal_row, col_idx, f"=SUM({letter}{data_start}:{letter}{data_end})")
    ws.cell(subtotal_row, 13, f"=IFERROR(P{subtotal_row}/D{subtotal_row},0)")

    totals: dict[str, Decimal] = {}
    for column in columns:
        if column.accumulable:
            totals[column.key] = sum((_decimal(getattr(m, column.key, None)) for m in member_results), Decimal("0"))
    return SummaryBlockResult(start_row, header_row, data_start, data_end, subtotal_row, subtotal_row + 3, len(member_results), totals)


def write_batch_grand_total(ws, row: int, block_results: Sequence[SummaryBlockResult], columns: Sequence[SummaryColumn]) -> int:
    if not block_results:
        return row
    row += 1
    ws.cell(row, 2, "TOTAL GENERAL DEL LOTE")
    totals: dict[str, Decimal] = {}
    for block in block_results:
        for key, value in block.numeric_totals.items():
            totals[key] = totals.get(key, Decimal("0")) + value
    for col_idx, column in enumerate(columns, start=1):
        if column.key in totals:
            ws.cell(row, col_idx, totals[column.key])
    total_amount = totals.get("total_amount", Decimal("0"))
    total_net = totals.get("net_kg", Decimal("0"))
    ws.cell(row, 13, (total_amount / total_net) if total_net else Decimal("0"))
    return row + 1


def _style_batch_detail_blocks(ws, block_results: Sequence[SummaryBlockResult]) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    for block in block_results:
        if block.subtotal_row:
            _mark_total_row(ws, block.subtotal_row, "E2F0D9")
    for cell in ws[ws.max_row - 1 if ws.max_row > 1 and ws.cell(ws.max_row, 2).value is None else ws.max_row]:
        if ws.cell(cell.row, 2).value == "TOTAL GENERAL DEL LOTE":
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="C6E0B4")
    formats = {
        1: SUMMARY_INTEGER_FORMAT, 4: SUMMARY_INTEGER_FORMAT,
        5: SUMMARY_MONEY_FORMAT, 7: SUMMARY_MONEY_FORMAT, 8: SUMMARY_MONEY_FORMAT, 9: SUMMARY_MONEY_FORMAT,
        10: SUMMARY_MONEY_FORMAT, 11: SUMMARY_MONEY_FORMAT, 12: SUMMARY_MONEY_FORMAT, 16: SUMMARY_MONEY_FORMAT,
        6: PRICE_FORMAT, 13: PRICE_FORMAT,
        14: SUMMARY_PERCENT_FORMAT, 15: SUMMARY_PERCENT_FORMAT,
        19: PTS_KG_FORMAT, 20: PTS_KG_FORMAT, 21: PTS_KG_FORMAT,
    }
    for col_idx, fmt in formats.items():
        for cell in ws[get_column_letter(col_idx)]:
            cell.number_format = fmt


def export_batch_liquidation_summary(results: Sequence, failed_results: Sequence, output_path: Path, *, campaign: str, company: str, crop: str, execution_started_at: datetime, execution_finished_at: datetime) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen por remesa"
    headers = ["Id Remesa", "Remesa", "Fecha de pago", "Periodo desde", "Periodo hasta", "Categoría", "Tipo liquidación", "Nº entregas", "Nº socios", "Nº variedades", "Kilos netos", "Importe comercial", "Recolección", "Calidad", "Transporte", "GlobalGAP", "Cuota Ha", "Base imponible", "Importe IVA", "Importe retención", "Importe total", "Estado", "Carpeta de salida"]
    ws.append(headers)
    totals = {h: Decimal("0") for h in headers[10:21]}
    for item in results:
        rem = item.remittance
        calc = item.calculation_result.result if hasattr(item.calculation_result, "result") else item.calculation_result
        t = calc.totals
        row = [rem.remittance_id, rem.name, rem.payment_date, rem.period_from, rem.period_to, rem.category, rem.liquidation_type, item.delivery_count, item.member_count, getattr(calc, "variety_count", ""), _n(t.net_kg), _n(t.commercial_amount), _n(t.collection_amount), _n(t.quality_amount), _n(t.transport_amount), _n(t.globalgap_amount), _n(t.hectare_fee_amount), _n(t.taxable_base), _n(t.vat_amount), _n(t.withholding_amount), _n(t.total_amount), "SUCCESS", str(item.output_directory)]
        ws.append(row)
        for idx, h in enumerate(headers[10:21], start=11):
            if row[idx - 1] is not None:
                totals[h] += row[idx - 1]
    if results:
        ws.append(["TOTAL GENERAL", "", "", "", "", "", "", sum(r.delivery_count for r in results), sum(r.member_count for r in results), "", *[totals[h] for h in headers[10:21]], "", ""])
        _mark_total_row(ws, ws.max_row, "C6E0B4")
    _style_sheet(ws, money_cols=range(12, 22), integer_cols=(8, 9, 10, 11), date_cols=(3, 4, 5))

    detail = wb.create_sheet("Detalle acumulado")
    block_results: list[SummaryBlockResult] = []
    detail_row = 1
    for index, item in enumerate(results):
        if index:
            detail.row_breaks.append(Break(id=detail_row))
        rem = item.remittance
        calc = item.calculation_result.result if hasattr(item.calculation_result, "result") else item.calculation_result
        block = append_liquidation_summary_block(detail, detail_row, remittance=rem, calculation_result=calc)
        block_results.append(block)
        detail_row = block.next_row
    if block_results:
        detail_row = write_batch_grand_total(detail, detail_row, block_results, get_liquidation_summary_columns())
    apply_liquidation_summary_style(
        detail,
        max(detail.max_row, 1),
        header_rows=[block.header_row for block in block_results],
        data_start_row=1,
        freeze=False,
        autofilter=False,
    )
    _style_batch_detail_blocks(detail, block_results)
    detail.page_setup.orientation = "landscape"
    detail.page_setup.fitToWidth = 1
    detail.page_setup.fitToHeight = 0
    detail.sheet_properties.pageSetUpPr.fitToPage = True
    detail.page_margins.left = 0.25
    detail.page_margins.right = 0.25
    detail.page_margins.top = 0.5
    detail.page_margins.bottom = 0.5

    inc = wb.create_sheet("Incidencias")
    inc.append(["Id Remesa", "Remesa", "Fase", "Tipo de error", "Mensaje", "Fecha y hora", "Estado"])
    for fail in failed_results:
        inc.append([fail.remittance.remittance_id, fail.remittance.name, fail.phase, fail.error_type, fail.error_message, execution_finished_at, "ERROR"])
    for item in results:
        calc = item.calculation_result.result if hasattr(item.calculation_result, "result") else item.calculation_result
        for warning in getattr(calc, "warnings", ()):
            inc.append([item.remittance.remittance_id, item.remittance.name, "WARNING", "Warning", warning, execution_finished_at, "WARNING"])
    _style_sheet(inc, date_cols=(6,))

    params = wb.create_sheet("Parámetros")
    first_calc = None
    if results:
        first_calc = results[0].calculation_result.result if hasattr(results[0].calculation_result, "result") else results[0].calculation_result
    master = getattr(first_calc, "hectare_fee_master", None)
    params_rows = [("Fecha y hora de inicio", execution_started_at), ("Fecha y hora de fin", execution_finished_at), ("Campaña", campaign), ("Empresa", company), ("Cultivo", crop), ("Usuario", ""), ("Número de remesas seleccionadas", len(results) + len(failed_results)), ("Número correctas", len(results)), ("Número con error", len(failed_results)), ("IDs seleccionados", ", ".join(str(r.remittance.remittance_id) for r in results) + (", " if results and failed_results else "") + ", ".join(str(f.remittance.remittance_id) for f in failed_results)), ("Nombres de las remesas", ", ".join(r.remittance.name for r in results) + (", " if results and failed_results else "") + ", ".join(f.remittance.name for f in failed_results)), ("Tarifa Cuota Ha", getattr(master, "price_per_hectare", "")), ("Cultivos sujetos a Cuota Ha", ", ".join(getattr(master, "eligible_crops", ()))), ("Versión de la aplicación", "Remesas"), ("Ruta base de salida", str(output_path.parent))]
    for row in params_rows:
        params.append(row)
    _style_sheet(params, date_cols=(2,))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        raise PermissionError("No se ha podido guardar el resumen acumulado porque el archivo está abierto en Excel. Cierre el archivo y pulse Reintentar.")
    return output_path
