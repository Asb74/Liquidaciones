from __future__ import annotations
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from domain.utils import to_decimal


def _text_or_empty(value):
    return "" if value is None else str(value)


def _joined_text(values):
    return ", ".join(_text_or_empty(value) for value in (values or ()) if value is not None)

def export_hectare_fee_report(path, summaries, crop_details, surface_details, incidents, campaign=None, company=None):
    wb=Workbook(); wb.remove(wb.active)
    sheets=[("Resumen por boleta", ["Socio","Agricultor","Boleta","Campaña","Empresa","Superficie","Precio/ha","Cuota Ha","Entregas","Cultivos","Índice €/kg","Cuota aplicada","Cuota pendiente","Estado"]), ("Detalle por cultivo", ["Socio","Agricultor","Boleta","Cultivo","Número de entregas","Kilos","Porcentaje","Índice €/kg","Cuota aplicada"]), ("Detalle de superficie", ["Socio","Agricultor","Boleta","Cultivo superficie","Variedad","Polígono","Parcela","Recinto","Superficie","CHA","Incluida","Motivo exclusión"]), ("Incidencias", ["Tipo","Socio","Boleta","Detalle"]), ("Boletas revisadas", ["Socio","Agricultor","Boleta","CHA","Nº parcelas","Superficie total","Superficie válida","Superficie excluida","Años detectados","Estado","Motivos","Incidencias"]), ("Detalle parcelas", ["Socio","Boleta","Polígono","Parcela","Recinto","SupCul","Año","Antigüedad","Incluida","Motivo","Incidencia"]), ("Incidencias Cuota Ha", ["Tipo","Socio","Boleta","Parcela","Valor encontrado","Descripción","Impide cálculo (Sí/No)"])]
    for name,headers in sheets:
        ws=wb.create_sheet(name); ws.append(headers); ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{chr(64+len(headers))}1"
        for cell in ws[1]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="1F4E78")
    for s in summaries:
        wb.worksheets[0].append([s.member_id,s.member_name,s.boleta,s.campaign,s.company,s.surface_hectares,s.price_per_hectare,s.annual_fee,s.total_delivery_kg," / ".join(s.delivery_crops),s.rate_per_kg,s.applied_fee,s.pending_fee,s.status])
        key=(s.member_id,s.boleta,s.campaign,s.company)
        for c in crop_details.get(key,()): wb.worksheets[1].append([s.member_id,s.member_name,s.boleta,c.crop,c.delivery_count,c.kilograms,c.percentage,c.rate_per_kg,c.applied_fee])
        for d in surface_details.get(key,()): wb.worksheets[2].append([s.member_id,s.member_name,s.boleta,d.crop,d.variety,d.polygon,d.parcel,d.enclosure,to_decimal(d.surface),"Sí" if d.cha_active else "No","Sí" if d.included else "No",d.exclusion_reason])
        audit = getattr(s, "audit", None)
        if audit:
            wb.worksheets[4].append([s.member_id,s.member_name,s.boleta,"Sí" if audit.get("cha") else "No",to_decimal(audit.get("numero_parcelas")),to_decimal(audit.get("superficie_total")),to_decimal(audit.get("superficie_valida")),to_decimal(audit.get("superficie_excluida")),_joined_text(audit.get("anos_detectados")),_text_or_empty(audit.get("estado_boleta")),"; ".join(_text_or_empty(value) for value in (audit.get("motivos_exclusion") or ()) if value is not None),"; ".join(_text_or_empty(value) for value in (audit.get("incidencias") or ()) if value is not None)])
            for parcel in audit.get("parcelas") or ():
                reason = parcel.get("Motivo exclusión", "")
                surface = to_decimal(parcel.get("SupCul DParcela"))
                incident = "PARCELA_SUPERFICIE_CERO" if surface == 0 else ""
                wb.worksheets[5].append([s.member_id,s.boleta,_text_or_empty(parcel.get("Pol")),_text_or_empty(parcel.get("Par")),_text_or_empty(parcel.get("Rec")),surface,_text_or_empty(parcel.get("Año")),_text_or_empty(parcel.get("Antigüedad")),_text_or_empty(parcel.get("Incluida")),_text_or_empty(reason),incident])
                if reason:
                    wb.worksheets[6].append([reason,s.member_id,s.boleta,_text_or_empty(parcel.get("Par")),surface,reason,"Sí" if parcel.get("Incluida") != "Sí" else "No"])
            for incident in audit.get("incidencias") or ():
                wb.worksheets[6].append([incident,s.member_id,s.boleta,"", "", incident, "No"])
    for row in incidents: wb.worksheets[3].append(row)
    # Keep the requested context even if the report is empty, so it is
    # unambiguous which company/campaign produced this workbook.
    context = wb.create_sheet("Parámetros")
    context.append(["Informe", "Cuota por hectárea"])
    context.append(["Campaña", campaign if campaign is not None else (summaries[0].campaign if summaries else "")])
    context.append(["Empresa", company if company is not None else (summaries[0].company if summaries else "")])
    for cell in context[1]:
        cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="1F4E78")
    for ws in wb.worksheets:
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(45,max(12,max(len(str(c.value or "")) for c in col)+2))
        for row in ws.iter_rows(min_row=2):
            for c in row:
                if isinstance(c.value, Decimal): c.number_format='#,##0.00########'
    wb.save(path); return path
