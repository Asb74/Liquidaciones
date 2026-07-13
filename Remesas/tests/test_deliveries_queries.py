from __future__ import annotations
import csv, sqlite3, tempfile, unittest
from datetime import date
from pathlib import Path
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None
from data.db_connection import AppConfig, ReadOnlyDatabase
from data.deliveries_repository import DeliveriesRepository
from domain.models import DeliveryFilter, Period, WorkContext

class DeliveriesQueryTests(unittest.TestCase):
    def setUp(self):
        self.tmp=tempfile.TemporaryDirectory(); p=Path(self.tmp.name); self.fruta=p/'f.sqlite'; self.eepp=p/'e.sqlite'
        with sqlite3.connect(self.fruta) as c:
            c.execute('CREATE TABLE PesosFres(CAMPAÑA TEXT, EMPRESA TEXT, CULTIVO TEXT, Fcarga TEXT, Reg INTEGER, IdSocio INTEGER, Variedad TEXT, Categoria TEXT, Neto REAL, Albaran TEXT, Boleta TEXT, Plataforma TEXT, Liquidado INTEGER)')
            c.executemany('INSERT INTO PesosFres VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)',[('2026','01','NAR','2026-01-05 10:00',1,7,'V1','A',100,'A1','B1','P',0),('2026','01','NAR','2026-01-06',2,8,'V2','B',50,'A2','B2','P',1)])
        with sqlite3.connect(self.eepp) as c:
            c.execute('CREATE TABLE DSocio(IdSocio INTEGER, Nombre TEXT)'); c.execute('INSERT INTO DSocio VALUES(7,"Socio 7")')
        cfg=AppConfig(str(self.fruta),str(self.eepp),'app','TEST',1,1,str(p/'x.log'),'INFO'); self.repo=DeliveriesRepository(ReadOnlyDatabase(cfg).connect_fruta_with_eepp())
    def tearDown(self): self.tmp.cleanup()
    def test_period_query_and_summary(self):
        rows,summary,_,total=self.repo.fetch(DeliveryFilter(WorkContext('2026','01','NAR'),Period(date(2026,1,1),date(2026,1,31))))
        self.assertEqual(total,2); self.assertEqual(summary.kilos_netos,150); self.assertEqual(rows[0].nombre_socio,'Socio 7')
    def test_socio_filter(self):
        rows,_,_,total=self.repo.fetch(DeliveryFilter(WorkContext('2026','01','NAR'),Period(date(2026,1,1),date(2026,1,31)),socio='7'))
        self.assertEqual(total,1); self.assertEqual(rows[0].socio,7)
    def test_variety_filter(self):
        rows,_,_,total=self.repo.fetch(DeliveryFilter(WorkContext('2026','01','NAR'),Period(date(2026,1,1),date(2026,1,31)),varieties=['V2']))
        self.assertEqual(total,1); self.assertEqual(rows[0].variedad,'V2')
    def test_csv_export_primitives(self):
        csv_path=Path(self.tmp.name)/'x.csv'
        with csv_path.open('w',newline='',encoding='utf-8') as f: csv.writer(f).writerows([['a'],['b']])
        self.assertEqual(csv_path.read_text(encoding='utf-8').splitlines()[0],'a')
    @unittest.skipIf(Workbook is None, "openpyxl no instalado en el entorno")
    def test_excel_export_primitives(self):
        xlsx_path=Path(self.tmp.name)/'x.xlsx'
        wb=Workbook(); wb.active.append(['ok']); wb.save(xlsx_path)
        self.assertEqual(load_workbook(xlsx_path).active['A1'].value,'ok')
if __name__ == '__main__': unittest.main()
