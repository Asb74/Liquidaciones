from __future__ import annotations

import tempfile
import unittest
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

from domain.calculation_models import LiquidationHeader, LiquidationResult, LiquidationTotals, MemberLiquidation
from exporters.excel_exporter import SUMMARY_HEADERS, export_liquidation_summary


class ExcelSummaryExporterTests(unittest.TestCase):
    def _result(self, *, net=Decimal("105845"), transport=Decimal("0")):
        header = LiquidationHeader(
            remesa_id=1,
            remesa_name="Tango Semana 5 NORMAL",
            campana="2026",
            empresa="1",
            cultivo="DIRECTO",
            fecha_pago="",
            periodo_desde="",
            periodo_hasta="",
            tipo_liquidacion="NORMAL",
            categoria="",
            socio="0",
            variedades=["TANGO"],
            options={},
            prices={},
        )
        member = MemberLiquidation(
            member_id=1561,
            member_name="SUAREZ SANCHEZ, Mª DEL PILAR",
            variety="TANGO",
            delivery_count=1,
            net_deliveries=net,
            net_commercial=net,
            net_waste=Decimal("0"),
            net_rotten=Decimal("0"),
            grades=(),
            commercial_amount=Decimal("90849.35"),
            gross_amount=Decimal("90849.35"),
            collection_amount=Decimal("0"),
            transport_amount=transport,
            quality_amount=Decimal("0"),
            globalgap_amount=Decimal("1812.54"),
            hectare_fee_amount=Decimal("656.01"),
            taxable_base=Decimal("92005.88"),
            vat_rate=Decimal("12"),
            withholding_rate=Decimal("2"),
            total_amount=Decimal("100985.66"),
            commercial_average_price=Decimal("0.85832"),
            final_average_price=Decimal("0.95409"),
        )
        totals = LiquidationTotals(
            net_kg=net,
            commercial_amount=Decimal("90849.35"),
            gross_amount=Decimal("90849.35"),
            detected_collection_amount=Decimal("0"),
            collection_amount=Decimal("0"),
            detected_transport_amount=transport,
            transport_amount=transport,
            quality_amount=Decimal("0"),
            globalgap_amount=Decimal("1812.54"),
            hectare_fee_amount=Decimal("656.01"),
            taxable_base=Decimal("92005.88"),
            vat_amount=Decimal("11040.71"),
            withholding_amount=Decimal("1840.12"),
            total_amount=Decimal("100985.66"),
        )
        return LiquidationResult(header, (member,), totals, ())

    def _export_and_load(self, result=None):
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        path = Path(tmp.name) / "resumen_liquidaciones.xlsx"
        export_liquidation_summary(result or self._result(), path)
        return path, load_workbook(path, data_only=False)["Resumen"]

    def test_column_order_formulae_totals_and_formats(self):
        path, ws = self._export_and_load()
        self.assertEqual([ws.cell(1, col).value for col in range(1, 22)], SUMMARY_HEADERS)
        self.assertEqual(ws.max_column, 21)
        self.assertEqual(ws["S2"].value, "=IFERROR(166.386*G2/D2,0)")
        self.assertEqual(ws["T2"].value, "=IFERROR(166.386*K2/D2,0)")
        self.assertEqual(ws["U2"].value, "=IFERROR(166.386*J2/D2,0)")
        expected = (Decimal("166.386") * Decimal("1812.54") / Decimal("105845")).quantize(Decimal("0.01"), ROUND_HALF_UP)
        self.assertEqual(expected, Decimal("2.85"))
        self.assertEqual(ws["B3"].value, "TOTAL")
        self.assertEqual(ws["D3"].value, "=SUM(D2:D2)")
        self.assertEqual(ws["F3"].value, None)
        self.assertEqual(ws["M3"].value, None)
        self.assertEqual(ws["S3"].value, None)
        self.assertEqual(ws["D2"].number_format, "#,##0;-#,##0;-")
        self.assertEqual(ws["G2"].value, 0)
        self.assertNotEqual(ws["G2"].value, "-")
        self.assertEqual(ws["G2"].number_format, "#,##0.00;-#,##0.00;-")
        self.assertEqual(ws["S2"].number_format, "0.00;-0.00;-")
        self.assertEqual(ws["N2"].value, 12)
        self.assertEqual(ws["N2"].number_format, '0"%"')
        self.assertEqual(ws.freeze_panes, "A2")
        self.assertIsNotNone(ws.auto_filter.ref)
        self.assertEqual(path.name, "resumen_liquidaciones.xlsx")

    def test_zero_net_uses_iferror_formulae(self):
        _, ws = self._export_and_load(self._result(net=Decimal("0")))
        self.assertEqual(ws["S2"].value, "=IFERROR(166.386*G2/D2,0)")
        self.assertEqual(ws["T2"].value, "=IFERROR(166.386*K2/D2,0)")
        self.assertEqual(ws["U2"].value, "=IFERROR(166.386*J2/D2,0)")

    def test_negative_transport_sign_is_preserved_by_formula(self):
        _, ws = self._export_and_load(self._result(transport=Decimal("-10.50")))
        self.assertEqual(ws["J2"].value, Decimal("-10.50"))
        self.assertEqual(ws["U2"].value, "=IFERROR(166.386*J2/D2,0)")
        expected = Decimal("166.386") * Decimal("-10.50") / Decimal("105845")
        self.assertLess(expected, 0)


if __name__ == "__main__":
    unittest.main()
