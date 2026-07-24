from decimal import Decimal
import sqlite3

from openpyxl import load_workbook

from data.hectare_repository import HectareRepository
from domain.hectare_fee_master import HectareFeeMaster
from exporters.hectare_fee_report_excel_exporter import export_hectare_fee_report
from services.hectare_fee_report_service import HectareFeeBoletaSummary, HectareFeeReportService


class MasterRepository:
    def __init__(self, crops): self.crops = crops
    def load(self): return HectareFeeMaster(Decimal("195"), tuple(self.crops))


def _repository():
    conn = sqlite3.connect(":memory:")
    conn.execute("CREATE TABLE PesosFres (IdSocio INTEGER, Socio TEXT, CAMPAÑA TEXT, EMPRESA TEXT, Boleta TEXT, CULTIVO TEXT, Neto REAL, NetoPartida REAL)")
    conn.executemany("INSERT INTO PesosFres VALUES (?, ?, ?, ?, ?, ?, ?, ?)", [
        (1, "Ana", "2026", "1", "100", " CITRICOS ", 50000, 0),
        (1, "Ana", "2026", "1", "100", "KAKIS", 20000, 0),
        (2, "Beto", "2026", "1", "200", "KAKIS", 30000, 0),
        (3, "Cora", "2026", "1", "", "KAKIS", 10000, 0),
    ])
    return HectareRepository(conn)


def test_repository_filters_inactive_crops_before_boleta_grouping_and_incidents():
    repository = _repository()

    boletas = repository.list_fee_report_boletas("2026", "1", active_fee_crops=("CITRICOS",))
    deliveries = repository.get_boleta_deliveries(1, "100", "2026", "1", active_fee_crops=("CITRICOS",))
    incidents = repository.list_deliveries_without_valid_boleta("2026", "1", ("CITRICOS",))

    assert [(row[0], row[4]) for row in boletas] == [(1, "100")]
    assert [(row[1].strip(), row[4]) for row in deliveries] == [("CITRICOS", 50000)]
    assert incidents == []
    assert repository.last_fee_report_query_counts == {"rows_read": 3, "rows_excluded_inactive_crop": 2, "rows_included": 1}


def test_report_and_excel_exclude_kakis_and_recalculate_rate(tmp_path):
    repository = _repository()
    service = HectareFeeReportService(repository, MasterRepository(("CITRICOS",)))
    # The surface query is outside the delivery filtering concern; provide one
    # valid parcel so this assertion also exercises fee allocation.
    repository.get_boleta_surface_details = lambda *_args: [
        ({"Cultivo": "CITRICOS", "CHA activo": "Sí"}, True, "", (None, None, None, None, None, None, None, None, None, Decimal("2")))
    ]

    summaries, crop_details, surface_details, incidents = service.build_report("2026", "1")

    assert len(summaries) == 1
    summary = summaries[0]
    assert summary.boleta == "100"
    assert summary.delivery_crops == ("CITRICOS",)
    assert summary.total_delivery_kg == Decimal("50000")
    assert summary.rate_per_kg == Decimal("0.0078")
    assert crop_details[(1, "100", "2026", "1")][0].applied_fee == Decimal("390.00")
    assert not incidents

    path = tmp_path / "informe.xlsx"
    export_hectare_fee_report(path, summaries, crop_details, surface_details, incidents)
    from openpyxl import load_workbook
    workbook = load_workbook(path, data_only=True)
    assert all("KAKIS" not in str(cell.value) for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row)


def test_master_change_is_read_on_each_report_without_a_stale_cache():
    repository = _repository()
    master = MasterRepository(("CITRICOS",))
    service = HectareFeeReportService(repository, master)
    repository.get_boleta_surface_details = lambda *_args: []

    assert [summary.boleta for summary in service.build_report("2026", "1")[0]] == ["100"]
    master.crops = ("KAKIS",)
    assert [summary.boleta for summary in service.build_report("2026", "1")[0]] == ["100", "200"]


def test_excel_exports_zero_and_optional_audit_values_without_changing_fee_amounts(tmp_path):
    normal = HectareFeeBoletaSummary(
        1, "Ana", "2026", "1", "100", Decimal("2"), Decimal("195"),
        Decimal("390"), Decimal("50000"), ("CITRICOS",), Decimal("0.0078"),
        Decimal("390.00"), Decimal("0.00"), "APLICADA",
    )
    cha_zero = HectareFeeBoletaSummary(
        2, "Beto", "2026", "1", "200", Decimal("0"), Decimal("195"),
        Decimal("0"), Decimal("0"), (), None, Decimal("0"), Decimal("0"),
        "SIN SUPERFICIE VÁLIDA", audit={
            "cha": False,
            "numero_parcelas": "",
            "superficie_total": None,
            "superficie_valida": None,
            "superficie_excluida": None,
            "anos_detectados": ("2018", None, "2020"),
            "estado_boleta": None,
            "motivos_exclusion": (None, "CHA_INACTIVA"),
            "incidencias": ("PARCELA_SUPERFICIE_CERO",),
            "parcelas": ({
                "Pol": None, "Par": None, "Rec": None,
                "SupCul DParcela": Decimal("0"), "Año": None,
                "Antigüedad": None, "Incluida": None,
                "Motivo exclusión": "PARCELA_SUPERFICIE_CERO",
            },),
        },
    )

    path = tmp_path / "cuota_ha.xlsx"
    export_hectare_fee_report(path, (normal, cha_zero), {}, {}, ())

    workbook = load_workbook(path, data_only=True)
    reviewed = workbook["Boletas revisadas"]
    parcels = workbook["Detalle parcelas"]
    incidents = workbook["Incidencias Cuota Ha"]

    assert reviewed.max_row == 2
    assert reviewed.cell(2, 4).value == "No"
    assert [reviewed.cell(2, column).value for column in (5, 6, 7, 8)] == [0, 0, 0, 0]
    assert reviewed.cell(2, 9).value == "2018, 2020"
    assert reviewed.cell(2, 10).value is None
    assert parcels.cell(2, 6).value == 0
    assert parcels.cell(2, 7).value is None
    assert parcels.cell(2, 8).value is None
    assert parcels.cell(2, 11).value == "PARCELA_SUPERFICIE_CERO"
    assert incidents.cell(2, 5).value == 0
    assert workbook["Resumen por boleta"].cell(2, 8).value == 390
