from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from exporters.batch_liquidation_excel_exporter import export_batch_liquidation_summary
from services.batch_remittance_service import BatchRemittanceService, FailedRemittanceResult, SelectedRemittance, SingleRemittanceBatchResult


def remittance(remittance_id=2204, name="BLANCA TEMPRANA sem 1"):
    return SelectedRemittance(remittance_id, name, date(2026, 7, 16), date(2026, 1, 1), date(2026, 1, 7), "1", "NORMAL", "2026", "1", "CITRICOS")


def calc(remittance_id=2204, member_id=1):
    # The persisted commercial average is deliberately different: the exporter
    # must derive P. Comer. from the same gross/net values it writes.
    member = SimpleNamespace(member_id=member_id, member_name="Socio", variety="NAVEL", net_kg=Decimal("100"), gross_amount=Decimal("50"), commercial_average_price=Decimal("0.7"), collection_amount=Decimal("1"), hectare_fee_amount=Decimal("2"), quality_amount=Decimal("3"), transport_amount=Decimal("4"), globalgap_amount=Decimal("5"), taxable_base=Decimal("35"), final_average_price=Decimal("0.35"), vat_rate=Decimal("12"), vat_amount=Decimal("4.2"), withholding_rate=Decimal("2"), withholding_amount=Decimal("0.7"), total_amount=Decimal("38.5"))
    totals = SimpleNamespace(net_kg=Decimal("100"), commercial_amount=Decimal("50"), collection_amount=Decimal("1"), quality_amount=Decimal("3"), transport_amount=Decimal("4"), globalgap_amount=Decimal("5"), hectare_fee_amount=Decimal("2"), taxable_base=Decimal("35"), vat_amount=Decimal("4.2"), withholding_amount=Decimal("0.7"), total_amount=Decimal("38.5"))
    header = SimpleNamespace(remesa_name=f"Remesa {remittance_id}")
    result = SimpleNamespace(member_results=(member,), totals=totals, header=header, warnings=(), variety_count=1, hectare_fee_master=SimpleNamespace(price_per_hectare=Decimal("195"), eligible_crops=("CITRICOS",)))
    return SimpleNamespace(result=result, member_count=1, delivery_count=2)


def batch_result(r):
    return SingleRemittanceBatchResult(r, calc(r.remittance_id), 1, 2, Path("/tmp") / str(r.remittance_id), ())


def test_batch_continues_after_single_remittance_failure(tmp_path):
    rems = [remittance(1), remittance(2), remittance(3)]
    def processor(r, cb):
        if r.remittance_id == 2:
            raise RuntimeError("boom")
        return batch_result(r)
    service = BatchRemittanceService(single_processor=processor, output_base=tmp_path, exporter=lambda *args, **kwargs: kwargs.get("output_path", args[2]) if False else args[2], log_dir=tmp_path / "logs")
    result = service.process(rems)
    assert result.remittances_completed == 2
    assert result.remittances_failed == 1
    assert [r.remittance.remittance_id for r in result.successful_results] == [1, 3]


def test_batch_cancellation_stops_before_next_and_exports_partial(tmp_path):
    rems = [remittance(1), remittance(2)]
    calls = []
    def processor(r, cb):
        calls.append(r.remittance_id)
        return batch_result(r)
    service = BatchRemittanceService(single_processor=processor, output_base=tmp_path, exporter=lambda *args, **kwargs: args[2], should_cancel=lambda: bool(calls), log_dir=tmp_path / "logs")
    result = service.process(rems)
    assert calls == [1]
    assert result.cancelled is True
    assert result.aggregate_excel_path.parent == tmp_path / "2026" / "CITRICOS"


def test_exporter_creates_consolidated_business_sheets_and_partial_state(tmp_path):
    output = tmp_path / "batch.xlsx"
    failed = [FailedRemittanceResult(remittance(2206), "CALCULATING", "ValueError", "bad data")]
    export_batch_liquidation_summary([batch_result(remittance(2204)), batch_result(remittance(2205))], failed, output, campaign="2026", company="1", crop="CITRICOS", execution_started_at=datetime.now(), execution_finished_at=datetime.now())
    wb = load_workbook(output, data_only=False)
    assert wb.sheetnames == ["Resumen general", "Detalle por socio", "Detalle calibres", "Advertencias", "Remesas fallidas"]
    assert wb["Resumen general"].cell(5, 1).value == "TOTAL GENERAL"
    assert wb["Resumen general"].cell(5, 32).value == "Parcial"
    assert wb["Remesas fallidas"].cell(2, 2).value == 2206


def test_same_member_is_kept_separate_by_remittance(tmp_path):
    output = tmp_path / "same-member.xlsx"
    export_batch_liquidation_summary([batch_result(remittance(1)), batch_result(remittance(2))], [], output, execution_started_at=datetime.now(), execution_finished_at=datetime.now())
    ws = load_workbook(output)["Detalle por socio"]
    assert [(ws.cell(r, 2).value, ws.cell(r, 7).value, ws.cell(r, 9).value) for r in (2, 3)] == [(1, 1, "NAVEL"), (2, 1, "NAVEL")]


def test_none_totals_remain_blank_and_are_pending(tmp_path):
    r = remittance(9)
    totals = SimpleNamespace(net_kg=None, commercial_amount=None, gross_amount=None, detected_collection_amount=None, collection_amount=None, detected_transport_amount=None, transport_amount=None, quality_amount=None, globalgap_amount=None, hectare_fee_amount=None, taxable_base=None, vat_amount=None, withholding_amount=None, total_amount=None)
    calculation = SimpleNamespace(result=SimpleNamespace(member_results=(), totals=totals, warnings=(), variety_count=0))
    item = SingleRemittanceBatchResult(r, calculation, 0, 0, tmp_path, ())
    output = tmp_path / "pending.xlsx"
    export_batch_liquidation_summary([item], [], output)
    ws = load_workbook(output)["Resumen general"]
    assert ws.cell(2, 15).value is None
    assert ws.cell(2, 32).value == "Pendiente"
    assert ws.cell(3, 15).value is None
