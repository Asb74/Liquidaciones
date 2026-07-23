from __future__ import annotations

import csv
import hashlib
import logging
import os
from datetime import datetime, timezone
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, simpledialog, ttk

from data.db_connection import ReadOnlyDatabase, load_config, setup_logging
from data.deliveries_repository import DeliveriesRepository
from data.metadata_repository import MetadataRepository
from data.variety_repository import VarietyRepository
from data.hectare_fee_master_repository import HectareFeeCropRepository
from data.group_benchmark_repository import GroupBenchmarkRepository
from data.remesas_repository import RemesasRepository
from domain.models import DeliveryFilter, Period, Remesa
from domain.document_models import LiquidationDocumentMode
from domain.varieties import STATUS_EMPTY_GROUP, STATUS_GROUP, STATUS_NOT_FOUND, STATUS_VARIETY, normalize_variety_text
from domain.hectare_fee_master import HectareFeeMasterRepository
from domain.validators import parse_user_date, validate_context, validate_period
from domain.utils import format_currency_es, format_decimal_es, format_display_date, format_integer_es, format_percentage_es, format_price_es, parse_yes_no, safe_path_part
from services.context_service import ContextService
from services.deliveries_service import DeliveriesService
from services.remesas_service import RemesasService
from services.calculation_service import CalculationService
from services.hectare_fee_master_service import HectareFeeMasterService
from services.local_database_sync_service import LocalDatabaseSyncService
from services.variety_group_service import VarietyGroupService
from services.group_benchmark_service import GroupBenchmarkService
from ui.context_panel import ContextPanel
from ui.deliveries_panel import COLUMNS, DeliveriesPanel
from ui.remesa_panel import RemesaPanel
from ui.summary_panel import SummaryPanel
from ui.hectare_fee_master_dialog import HectareFeeMasterDialog
from exporters.excel_exporter import export_liquidation_summary
from exporters.batch_liquidation_excel_exporter import export_batch_liquidation_summary
from services.batch_remittance_service import BatchProgress, BatchRemittanceService, SelectedRemittance, SingleRemittanceBatchResult
from exporters.file_lock import FileLockedError
from exporters.hectare_fee_auditor import export_hectare_fee_audit
from presentation.premium_liquidation_view_model import from_member_liquidation
from presentation.liquidation_document_snapshot import SCHEMA_VERSION, dump as dump_document_snapshot
from data.persistence.database import PersistenceDatabase
from data.persistence.master_repository import LiquidationMasterRepository
from services.liquidation_persistence_service import LiquidationPersistenceService
from data.persistence.liquidation_repository import LiquidationRepository
from services.document_generation_service import DocumentGenerationOptions, DocumentGenerationService
from services.liquidation_history_service import LiquidationHistoryService
from services.liquidation_modification_service import LiquidationModificationService
from services.liquidation_csv_export_service import LiquidationCsvExportService
from services.pdf_preview_service import PdfPreviewService
from ui.persistence_result_dialog import PersistenceResultDialog
from ui.liquidation_history_dialog import LiquidationHistoryDialog
from ui.batch_persistence_preview_dialog import BatchPersistencePreviewDialog
from ui.liquidation_prefix_master_dialog import LiquidationPrefixMasterDialog
from ui.liquidation_split_master_dialog import LiquidationSplitMasterDialog
import json

logger = logging.getLogger(__name__)
from exporters.pdf_exporter import export_member_pdf
from exporters.premium_pdf_exporter import LOCKED_PDF_MESSAGE, generate_liquidation_pdf, premium_member_filename

class RemesasFrame(ttk.Frame):
    def __init__(self, master, config_path: str | None = None):
        super().__init__(master, padding=8)
        self.config=load_config(config_path); setup_logging(self.config)
        self.preview_service=PdfPreviewService()
        self.db=ReadOnlyDatabase(self.config); self.conn=None; self.current_persisted_batch_ids=(); self.current_generated_documents=(); self.current_persistence_status=None; self.selected_history_batch_ids=(); self.batch_cancel_requested=False; self.batch_running=False; self.variety_service=None; self.selected_source_items=[]; self.variety_resolutions=[]; self.summary=None; self.current_remesa=None; self.current_calculation=None; self.current_deliveries=[]; self.calculation_valid=False; self.current_calculation_persisted=False; self.current_batch_result=None; self.current_batch_preview=None; self.current_batch_persisted=False; self.current_batch_save_result=None; self.sync_results=[]; self.current_group_benchmarks={}; self.master_repository=HectareFeeMasterRepository(); self.hectare_master_service=HectareFeeMasterService(self.master_repository); self.active_master=self.hectare_master_service.load_master()
        self._build(); self._connect()

    def _build(self):
        self.db_status_text=tk.StringVar(value="Preparando bases locales...")
        ttk.Label(self,textvariable=self.db_status_text).grid(row=0,column=0,columnspan=3,sticky="ew")
        self.context_panel=ContextPanel(self,self.config,self._context_changed); self.context_panel.grid(row=1,column=0,columnspan=3,sticky="ew")
        self.remesa_panel=RemesaPanel(self); self.remesa_panel.grid(row=2,column=0,sticky="nsew",padx=4,pady=4)
        self._build_varieties(); self.var_frame.grid(row=2,column=1,sticky="nsew",padx=4,pady=4)
        self.summary_panel=SummaryPanel(self); self.summary_panel.grid(row=2,column=2,rowspan=2,sticky="nsew",padx=4,pady=4)
        self.deliveries_panel=DeliveriesPanel(self); self.deliveries_panel.grid(row=3,column=0,columnspan=2,sticky="nsew",padx=4,pady=4)
        self._build_hectare_config_info(); self.hectare_info.grid(row=4,column=0,columnspan=3,sticky="ew",pady=3)
        self._build_buttons(); self.buttons.grid(row=5,column=0,columnspan=3,sticky="ew")
        self.status=tk.StringVar(value="Listo"); ttk.Label(self,textvariable=self.status).grid(row=6,column=0,columnspan=3,sticky="ew")
        self.rowconfigure(3,weight=1); self.columnconfigure(1,weight=1)

    def _build_varieties(self):
        self.var_frame=ttk.LabelFrame(self,text="Variedades y configuración")
        self.available=tk.Listbox(self.var_frame,height=7,exportselection=False); self.selected=tk.Listbox(self.var_frame,height=7,exportselection=False)
        ttk.Label(self.var_frame,text="Disponibles").grid(row=0,column=0); ttk.Label(self.var_frame,text="Seleccionadas").grid(row=0,column=2)
        self.available.grid(row=1,column=0,rowspan=4,sticky="nsew"); self.selected.grid(row=1,column=2,rowspan=4,sticky="nsew")
        ttk.Button(self.var_frame,text="Añadir",command=self._add_var).grid(row=1,column=1); ttk.Button(self.var_frame,text="Quitar",command=self._remove_var).grid(row=2,column=1); ttk.Button(self.var_frame,text="Añadir todas",command=self._add_all_var).grid(row=3,column=1); ttk.Button(self.var_frame,text="Limpiar",command=self._clear_selected_varieties).grid(row=4,column=1)
        self.apply_collection_var=tk.BooleanVar(value=False); self.apply_transport_var=tk.BooleanVar(value=False); self.apply_quality_var=tk.BooleanVar(value=False); self.apply_globalgap_var=tk.BooleanVar(value=False); self.apply_hectare_fee_var=tk.BooleanVar(value=False); self.apply_precalibrated_var=tk.BooleanVar(value=False)
        self.option_vars={"Aplica recolección":self.apply_collection_var,"Aplica transporte":self.apply_transport_var,"Aplica calidad":self.apply_quality_var,"Aplica GlobalGAP":self.apply_globalgap_var,"Aplica cuota por hectárea":self.apply_hectare_fee_var,"Liquida precalibrado":self.apply_precalibrated_var}
        for i,(name,var) in enumerate(self.option_vars.items()):
            ttk.Checkbutton(self.var_frame,text=name,variable=var,command=self._invalidate_calculation).grid(row=5+i//2,column=i%2,sticky="w",columnspan=1)
        self.prices=ttk.LabelFrame(self.var_frame,text="Precios") ; self.prices.grid(row=8,column=0,columnspan=3,sticky="ew")
        self.resolved_selection_text=tk.StringVar(value="Selección resuelta: ")
        ttk.Label(self.var_frame,textvariable=self.resolved_selection_text,wraplength=520,justify="left").grid(row=7,column=0,columnspan=3,sticky="ew",pady=(4,0))
        self.price_vars={f"P{i}":tk.StringVar(value="Pendiente") for i in range(12)} | {"PDESTRIO":tk.StringVar(value="Pendiente"),"PDMESA":tk.StringVar(value="Pendiente"),"PPODRIDO":tk.StringVar(value="Pendiente")}
        for i,(k,v) in enumerate(self.price_vars.items()): ttk.Label(self.prices,text=(f"Calibre {k[1:]}" if k.startswith('P') and k[1:].isdigit() else k)).grid(row=i//4,column=(i%4)*2); ttk.Label(self.prices,textvariable=v).grid(row=i//4,column=(i%4)*2+1)

    def _build_buttons(self):
        self.buttons=ttk.Frame(self)
        actions=[("Nueva remesa",self._new_remittance),("Cargar remesa existente",self._load_remesa),("Buscar entregas",self._search),("Exportar entregas a CSV",self._export_csv),("Exportar entregas a Excel",self._export_excel),("Vista previa",self._preview),("Cerrar",self.close_application)]
        self.action_buttons={}
        for i,(text,cmd) in enumerate(actions):
            b=ttk.Button(self.buttons,text=text,command=cmd); b.grid(row=0,column=i,padx=2); self.action_buttons[text]=b
        for i,(text,cmd) in enumerate([("Calcular liquidación",self._calculate),("Exportar resumen de liquidación",self._export_liquidation_excel),("Vista previa PDF",self._export_premium_pdf), ("Informe interno",self._export_liquidation_pdf),("Revisar lote",self._review_batch),("Guardar liquidaciones",self._save_liquidations),("Anular liquidación",self._void_liquidation)]):
            b=ttk.Button(self.buttons,text=text,command=cmd,state="disabled"); b.grid(row=1,column=i,padx=2,pady=2); self.action_buttons[text]=b
        self.persistence_status_text=tk.StringVar(value="Sin cálculo pendiente.")
        self.persistence_status_label=ttk.Label(self.buttons,textvariable=self.persistence_status_text); self.persistence_status_label.grid(row=2,column=0,columnspan=9,sticky="w")

    def _build_hectare_config_info(self):
        self.hectare_info=ttk.LabelFrame(self,text="Configuración cuota Ha")
        self.hectare_config_text=tk.StringVar(value="")
        ttk.Label(self.hectare_info,textvariable=self.hectare_config_text,wraplength=1100).pack(side="left",fill="x",expand=True,padx=6)
        self._refresh_hectare_config_label()

    def _refresh_hectare_config_label(self):
        m=self.active_master
        self.hectare_config_text.set(f"{str(m.price_per_hectare).replace('.', ',')} €/ha | Cultivos sujetos a Cuota Ha: {', '.join(m.eligible_crops)}")

    def open_hectare_fee_master(self):
        self._open_hectare_fee_master()

    def open_liquidation_history(self):
        if not getattr(self,"persistence_enabled",False):
            messagebox.showwarning("Historial","La persistencia local no está habilitada.")
            return
        LiquidationHistoryDialog(self.winfo_toplevel(),self.history_service)

    def open_liquidation_prefix_master(self):
        try:
            LiquidationPrefixMasterDialog(self.winfo_toplevel(),self.liquidation_master_repository,on_saved=self._persistence_master_saved)
            logger.info("[MasterDialog]\ntype=PREFIX\nopened=true")
        except Exception:
            logger.exception("No se ha podido abrir el maestro de prefijos")
            messagebox.showerror("Maestro de prefijos","No se ha podido abrir el maestro de prefijos de liquidación.")

    def open_liquidation_split_master(self):
        try:
            LiquidationSplitMasterDialog(
                self.winfo_toplevel(), self.liquidation_master_repository,
                member_name_lookup=self.persistence_service.legacy.member_name,
                on_saved=self._persistence_master_saved,
            )
            logger.info("[MasterDialog]\ntype=SPLIT\nopened=true")
        except Exception:
            logger.exception("No se ha podido abrir el maestro de divisiones")
            messagebox.showerror("Maestro de divisiones","No se ha podido abrir el maestro de división de liquidaciones.")

    def _persistence_master_saved(self):
        if self.current_batch_result is not None:
            self.current_batch_preview=None
            self.persistence_status_text.set("Las reglas de persistencia han cambiado; la vista previa se regenerará al guardar.")
        self._refresh_action_states()

    def show_about(self):
        self._show_about()

    def _show_about(self):
        messagebox.showinfo("Acerca de", "Liquidaciones - Remesas\nModo prueba SQLite\nVersión actual del módulo")

    def _open_hectare_fee_master(self):
        try:
            HectareFeeMasterDialog(self.winfo_toplevel(), self.hectare_master_service, on_saved=self._hectare_master_saved)
        except Exception:
            logger.exception("No se ha podido abrir el maestro de cuota por hectárea")
            messagebox.showerror("Maestro cuota Ha", "No se ha podido abrir el maestro de cuota por hectárea.")

    def _hectare_master_saved(self):
        self.active_master = self.hectare_master_service.load_master()
        self._refresh_hectare_config_label()
        self._invalidate_master_changed()

    def _connect(self):
        try:
            self.conn=self.db.connect_fruta_with_eepp(); self.meta=ContextService(MetadataRepository(self.conn)); self.variety_service=VarietyGroupService(VarietyRepository(self.conn)); self.deliveries=DeliveriesService(DeliveriesRepository(self.conn)); self.remesas=RemesasService(RemesasRepository(self.conn))
            self.persistence_enabled=False
            if self.config.persistence_enabled:
                aliases_path=Path(__file__).resolve().parents[1]/"config"/"crop_aliases.json"
                aliases=json.loads(aliases_path.read_text(encoding="utf-8")) if aliases_path.exists() else {}
                self.persistence_service=LiquidationPersistenceService(PersistenceDatabase(self.config.persistence_database_path),self.conn,crop_aliases=aliases)
                output_root=Path("C:/Liquidaciones/salidas/remesas") if Path("C:/").exists() else Path.cwd().parent/"salidas"/"remesas"
                self.liquidation_repository=LiquidationRepository(self.persistence_service.database)
                self.document_service=DocumentGenerationService(self.liquidation_repository, output_root)
                self.modification_service=LiquidationModificationService(self.persistence_service)
                self.csv_export_service=LiquidationCsvExportService(self.liquidation_repository, self.persistence_service.legacy, output_root)
                self.history_service=LiquidationHistoryService(self.liquidation_repository,self.document_service,self.modification_service,self.csv_export_service)
                self.liquidation_master_repository=LiquidationMasterRepository(self.persistence_service.database)
                self.persistence_service.import_legacy_split_rules(); self.persistence_enabled=True
            self.context_panel.campaña_cb["values"]=self.meta.campaigns(); self.context_panel.set_status(self.db.status()); self.hectare_master_service=HectareFeeMasterService(self.master_repository, HectareFeeCropRepository(self.conn)); self.calculations=CalculationService(self.conn, self.config); self._refresh_database_status(); self._refresh_action_states()
        except Exception as exc:
            logger.exception("No se ha podido abrir la copia local de las bases SQLite")
            messagebox.showerror("Error", "No se han podido preparar las bases de datos.\n\nDetalle:\nNo existe una copia local válida para abrir en modo lectura.\n\nRevise la conexión de red o utilice la última copia local disponible.")

    def _save_liquidations(self):
        if self.current_batch_result is not None:
            return self._save_batch_liquidations()
        return self._save_individual_liquidation()

    def _save_individual_liquidation(self):
        try:
            if not self.calculation_valid or not self.current_calculation or not self.current_calculation.result:
                raise ValueError("Debe calcular y revisar una liquidación válida")
            preview=self.persistence_service.prepare_preview(self.current_calculation.result)
            win=tk.Toplevel(self); win.title("Vista previa de guardado de liquidaciones"); win.transient(self.winfo_toplevel()); win.grab_set(); win.geometry("1150x650")
            nb=ttk.Notebook(win); nb.pack(fill="both",expand=True,padx=8,pady=8)
            summary=ttk.Frame(nb); splits=ttk.Frame(nb); final=ttk.Frame(nb); warnings=ttk.Frame(nb)
            for tab,title in ((summary,"Resumen"),(splits,"Repartos"),(final,"Líneas finales"),(warnings,"Advertencias")): nb.add(tab,text=title)
            original=self.current_calculation.result
            summary_text=(f"Remesa: {preview.header.remesa_name}\nCampaña: {preview.header.campana} | Empresa: {preview.header.empresa} | Cultivo: {preview.header.cultivo}\nFecha de pago: {preview.header.fecha_pago}\n"
                          f"Líneas originales/finales: {preview.original_line_count}/{len(preview.lines)}\nSocios originales/destinatarios: {len(set(x.member_id for x in original.member_results))}/{len(set(x.recipient_member_id for x in preview.lines))}\n"
                          f"Neto original/final: {original.totals.net_kg}/{sum((x.net_kg for x in preview.lines))}\nBase original/final: {original.totals.taxable_base}/{sum((x.taxable_base for x in preview.lines))}\nTotal fiscal original/final: {original.totals.total_amount}/{sum((x.total_amount for x in preview.lines))}\n\nEl total fiscal final puede variar si los destinatarios tienen regímenes distintos.")
            ttk.Label(summary,text=summary_text,justify="left").pack(anchor="nw",padx=12,pady=12)
            cols=("origen","destino","variedad","factor","neto","base","iva","ret","total")
            tree=ttk.Treeview(final,columns=cols,show="headings"); [tree.heading(c,text=c.title()) for c in cols]; tree.pack(fill="both",expand=True)
            for x in preview.lines: tree.insert("","end",values=(x.source_member_id,x.recipient_member_id,x.variety,x.split_factor,x.net_kg,x.taxable_base,x.vat_rate,x.withholding_rate,x.total_amount))
            ttk.Label(splits,text="\n".join(f"{x.source_member_id} → {x.recipient_member_id}: {x.split_factor} ({x.split_type or 'SIN DIVISIÓN'})" for x in preview.lines),justify="left").pack(anchor="nw",padx=12,pady=12)
            ttk.Label(warnings,text="\n".join(preview.warnings) or "Sin advertencias",justify="left").pack(anchor="nw",padx=12,pady=12)
            result={"confirm":False}
            def confirm(): result["confirm"]=True; win.destroy()
            buttons=ttk.Frame(win); buttons.pack(fill="x",padx=8,pady=8); ttk.Button(buttons,text="Confirmar y guardar",command=confirm).pack(side="right"); ttk.Button(buttons,text="Cancelar",command=win.destroy).pack(side="right",padx=6)
            ttk.Label(buttons,text="No se escribirá en Access. Los PDF definitivos se generan después del commit.").pack(side="left")
            win.wait_window()
            if not result["confirm"]: return
            batch=self.persistence_service.save(preview)
            # Persist the exact presentation model used by the draft. Regeneration
            # therefore never queries Access nor rebuilds a reduced PDF payload.
            for member in self._premium_members():
                if member.member_id == 0:
                    continue
                self.liquidation_repository.save_document_snapshot(
                    batch_id=batch.batch_id, recipient_member_id=member.member_id,
                    payload_json=dump_document_snapshot(from_member_liquidation(self.current_calculation.result.header, member, group_benchmark=self._benchmark_for_member(member))),
                    schema_version=SCHEMA_VERSION, calculation_fingerprint=preview.fingerprint,
                    created_at=datetime.now(timezone.utc).isoformat(),
                )
            documents=self.document_service.generate_for_batch(batch.batch_id,options=DocumentGenerationOptions())
            self.current_calculation_persisted=True
            self.current_persisted_batch_ids=(batch.batch_id,)
            self.current_generated_documents=documents.generated_documents+documents.failed_documents
            self.current_persistence_status="ACTIVE"
            logger.info("[PostPersistenceState]\nbatch_ids=%s\nactive_batches=%s\ndocuments=%s\ncan_void=true\ncan_view=%s",self.current_persisted_batch_ids,self.current_persisted_batch_ids,len(self.current_generated_documents),bool(documents.generated_documents))
            PersistenceResultDialog(self,self.history_service,self.current_persisted_batch_ids,self.current_generated_documents,len(batch.liquidations))
            self.status.set(f"Persistencia completada: batch {batch.batch_id}")
            self._refresh_action_states()
        except Exception as exc:
            logger.exception("No se pudieron guardar las liquidaciones"); messagebox.showerror("Guardar liquidaciones",str(exc))

    def _ensure_batch_preview(self):
        if self.current_batch_result is not None and self.current_batch_preview is None:
            self.current_batch_preview=self.persistence_service.prepare_batch_preview(self.current_batch_result)
            messagebox.showinfo("Vista previa actualizada","Las reglas de división o prefijos han cambiado. Se ha actualizado la vista previa de guardado.")
        return self.current_batch_preview

    def _review_batch(self):
        preview=self._ensure_batch_preview()
        if preview: BatchPersistencePreviewDialog(self.winfo_toplevel(),preview).show()

    def _save_batch_liquidations(self):
        try:
            if self.current_batch_persisted: raise ValueError("El lote ya fue guardado correctamente")
            preview=self._ensure_batch_preview()
            if not preview or not preview.valid: raise ValueError("El lote no contiene remesas guardables")
            valid=sum(x.valid for x in preview.remittances); total=valid+len(preview.excluded_remittances)
            dialog=BatchPersistencePreviewDialog(self.winfo_toplevel(),preview,allow_confirm=True)
            if not dialog.show(): return False
            if not messagebox.askyesno("Confirmar guardado",f"Se van a guardar las liquidaciones definitivas de {valid} remesas.\n\nUna vez guardadas se generarán los PDFs definitivos por destinatario.\n\nEl Excel resumen seguirá mostrando las liquidaciones originales sin división.\n\nSe guardarán {valid} de {total} remesas.\n\n¿Desea continuar?"): return False
            result=self.persistence_service.save_batch(preview); self.current_batch_save_result=result; self.current_batch_persisted=result.failed==0
            batch_ids=[x.batch.batch_id for x in result.remittance_results if x.saved]
            documents=self.document_service.generate_for_batches(batch_ids,options=DocumentGenerationOptions(),progress_callback=self._document_progress,cancel_requested=lambda:self.batch_cancel_requested)
            self.current_persisted_batch_ids=tuple(batch_ids)
            self.current_generated_documents=tuple(d for r in documents.results for d in r.generated_documents+r.failed_documents)
            self.current_persistence_status="ACTIVE" if result.failed==0 else "PARTIAL"
            status="SUCCESS" if not result.failed else "PARTIAL"
            logger.info("[BatchPersistence]\nrequested=%s\nsaved=%s\nfailed=%s\nstatus=%s",result.requested,result.saved,result.failed,status)
            document_failures=sum(len(x.failed_documents) for x in documents.results)
            PersistenceResultDialog(self,self.history_service,batch_ids,self.current_generated_documents,sum(len(x.batch.liquidations) for x in result.remittance_results if x.saved))
            self._refresh_action_states(); return self.current_batch_persisted
        except Exception as exc:
            logger.exception("No se pudo guardar el lote"); messagebox.showerror("Guardar liquidaciones",str(exc)); return False

    def _void_liquidation(self):
        try:
            if self.current_persisted_batch_ids:
                active=[bid for bid in self.current_persisted_batch_ids if (self.liquidation_repository.get_batch(bid) and self.liquidation_repository.get_batch(bid)["status"]=="ACTIVE")]
                if not active: messagebox.showinfo("Anular liquidación","No hay batches activos."); return
                batch_id=active[0] if len(active)==1 else simpledialog.askstring("Anular liquidación","Batch ID a anular:\n"+"\n".join(active),parent=self)
                reason=simpledialog.askstring("Anular liquidación",f"Motivo obligatorio para anular el batch {batch_id}:",parent=self) if batch_id else None
                if reason and messagebox.askyesno("Confirmar anulación",f"¿Anular el batch {batch_id}?",parent=self): self.history_service.void_batch(batch_id,reason); self.current_persistence_status="VOIDED"; logger.info("[BatchVoided]\nbatch_id=%s\nreason=%s",batch_id,reason); self._refresh_action_states(); messagebox.showinfo("Anular liquidación","Liquidación anulada.")
                return
            if self.current_batch_save_result and self.current_batch_save_result.saved:
                choices=[x for x in self.current_batch_save_result.remittance_results if x.saved]
                listing="\n".join(f"{i+1}. Remesa {x.remittance.remittance_id} - batch {x.batch.batch_id}" for i,x in enumerate(choices))
                selected=simpledialog.askinteger("Anular liquidación",f"Seleccione el batch que desea anular:\n\n{listing}",minvalue=1,maxvalue=len(choices),parent=self)
                if selected is None:return
                batch_id=choices[selected-1].batch.batch_id
                reason=simpledialog.askstring("Anular liquidación",f"Motivo obligatorio para anular el batch {batch_id}:",parent=self)
                if reason and messagebox.askyesno("Confirmar anulación",f"¿Anular el batch {batch_id}?",parent=self): self.persistence_service.void_batch(batch_id,reason); messagebox.showinfo("Anular liquidación","Liquidación anulada.")
                return
            remesa_id=int(self.current_calculation.result.header.remesa_id)
            with self.persistence_service.database.connect() as conn:
                rows=conn.execute("SELECT batch_id,created_at FROM liquidation_batches WHERE remesa_id=? AND status='ACTIVE' ORDER BY created_at DESC",(remesa_id,)).fetchall()
            if not rows: messagebox.showinfo("Anular liquidación","No hay batches activos para esta remesa."); return
            batch_id=str(rows[0]["batch_id"]); reason=simpledialog.askstring("Anular liquidación",f"Motivo obligatorio para anular el batch {batch_id}:",parent=self)
            if not reason: return
            if messagebox.askyesno("Confirmar anulación",f"¿Anular el batch {batch_id}?",parent=self): self.persistence_service.void_batch(batch_id,reason); messagebox.showinfo("Anular liquidación","Liquidación anulada.")
        except Exception as exc: logger.exception("No se pudo anular"); messagebox.showerror("Anular liquidación",str(exc))

    def _document_progress(self, event):
        self.status.set(f"Generando documentos · {event.get('phase','')} · {event.get('path','')}")
        self.update_idletasks()

    def synchronize_local_databases(self, manual: bool = False) -> bool:
        if manual and self.calculation_valid and not messagebox.askyesno("Actualizar bases locales", "Existe un cálculo activo. La actualización limpiará los resultados actuales. ¿Continuar?"):
            return False
        try:
            service=LocalDatabaseSyncService(self.config, progress_callback=self.status.set if hasattr(self, "status") else None)
            self.sync_results=service.synchronize_all()
            errors=[r for r in self.sync_results if not (r.synchronized or r.used_local_fallback)]
            if errors:
                detail="\n".join(f"{r.database_name}: {r.error_message}" for r in errors)
                messagebox.showerror("Bases de datos", f"No se han podido preparar las bases de datos.\n\nDetalle:\n{detail}\n\nRevise la conexión de red o utilice la última copia local disponible.")
                return False
            if manual:
                self._reconnect_after_sync()
                messagebox.showinfo("Bases locales", self._fallback_warning_text() or "Bases locales actualizadas desde red.")
            self._refresh_database_status()
            return True
        except Exception as exc:
            logger.exception("Error de sincronización manual")
            messagebox.showerror("Bases de datos", f"No se han podido preparar las bases de datos.\n\nDetalle:\n{exc}\n\nRevise la conexión de red o utilice la última copia local disponible.")
            return False

    def _reconnect_after_sync(self):
        if self.conn is not None:
            self.conn.close()
        self._clear()
        self._connect()

    def open_data_folder(self):
        Path(self.config.local_database_dir).mkdir(parents=True, exist_ok=True)
        if hasattr(os, "startfile"):
            os.startfile(self.config.local_database_dir)
        else:
            messagebox.showinfo("Carpeta de datos", self.config.local_database_dir)

    def _fallback_warning_text(self) -> str:
        if not any(r.used_local_fallback for r in self.sync_results):
            return ""
        lines=["No se ha podido acceder a las bases de red.", "", "La aplicación utilizará la última copia local válida:"]
        for r in self.sync_results:
            if r.used_local_fallback:
                stamp=r.local_modified_at.strftime("%d/%m/%Y %H:%M") if r.local_modified_at else "fecha desconocida"
                lines.append(f"{r.database_name}: {stamp}")
        lines += ["", "Los datos pueden no estar actualizados."]
        return "\n".join(lines)

    def _refresh_database_status(self):
        if not hasattr(self, "db_status_text"):
            return
        status=self.db.status()
        parts=[]
        for key in ("DBfruta", "DBEEPPL"):
            parts.append(f"{key} local: {status.get(key, 'No accesible')}")
        if self.sync_results:
            parts.append(" | ".join(sorted({r.status for r in self.sync_results})))
        self.db_status_text.set("   ".join(parts))

    def _context_changed(self):
        if self._has_pending_batch() and not self._confirm_discard_pending_batch():
            return
        self.current_remesa=None; self.remesa_panel.load({}); self._clear_selected_varieties(invalidate=False); self.deliveries_panel.clear(); self.summary_panel.clear(); self.current_calculation=None; self.calculation_valid=False; self.current_deliveries=[]; self.current_group_benchmarks={}
        ctx=self.context_panel.context()
        try:
            if ctx.campana and not ctx.empresa: self.context_panel.empresa_cb["values"]=self.meta.empresas(ctx.campana)
            if ctx.campana and ctx.empresa and not ctx.cultivo: self.context_panel.cultivo_cb["values"]=self.meta.cultivos(ctx.campana,ctx.empresa)
            if ctx.campana and ctx.empresa and ctx.cultivo: self._load_varieties()
        except Exception as exc: messagebox.showerror("Error",str(exc))
        self._refresh_action_states()

    def _load_varieties(self):
        self.available.delete(0,"end"); ctx=self.context_panel.context()
        d=self.remesa_panel.data(); desde=hasta=None
        try:
            if d.get("desde") and d.get("hasta"):
                desde=parse_user_date(d["desde"]); hasta=parse_user_date(d["hasta"])
        except ValueError:
            pass
        values = self.variety_service.list_selection_options(ctx.cultivo) if self.variety_service else ()
        if not values:
            values = self.meta.variedades(ctx.campana,ctx.empresa,ctx.cultivo,desde,hasta)
        for v in values: self.available.insert("end",v)
    def _clear_selected_varieties(self, invalidate: bool = True):
        self.selected.delete(0,"end")
        self.selected_source_items=[]; self.variety_resolutions=[]; self._refresh_resolved_selection_label()
        if invalidate:
            self._invalidate_calculation()
    def _add_var(self):
        for i in self.available.curselection():
            self._add_source_variety(self.available.get(i))
        self._invalidate_calculation()

    def _add_all_var(self):
        self._clear_selected_varieties(invalidate=False)
        for v in self.available.get(0,"end"): self._add_source_variety(v)
        self._invalidate_calculation()

    def _remove_var(self):
        for i in reversed(self.selected.curselection()): self.selected.delete(i)
        self.selected_source_items=list(self.selected.get(0,"end")); self.variety_resolutions=[]; self._refresh_resolved_selection_label(); self._invalidate_calculation()

    def _add_source_variety(self, value: str, *, show_warning: bool = True):
        ctx=self.context_panel.context(); res=self.variety_service.resolve_selection(ctx.cultivo, value) if self.variety_service else None
        if res and res.status in {STATUS_NOT_FOUND, STATUS_EMPTY_GROUP}:
            if show_warning:
                messagebox.showwarning("Variedades", res.warnings[0] if res.warnings else f"No se pudo resolver la variedad o grupo “{value}”.")
            return
        values = res.varieties if res else (value,)
        if value not in self.selected_source_items: self.selected_source_items.append(value)
        if res: self.variety_resolutions.append(res)
        current={normalize_variety_text(v) for v in self.selected.get(0,"end")}
        for variety in values:
            if normalize_variety_text(variety) not in current:
                self.selected.insert("end", variety); current.add(normalize_variety_text(variety))
        self._refresh_resolved_selection_label()

    def _refresh_resolved_selection_label(self):
        lines=[]
        for res in getattr(self, "variety_resolutions", []):
            if res.status == STATUS_VARIETY and res.varieties:
                lines.append(f"variedad concreta → {res.varieties[0]}")
            elif res.status == STATUS_GROUP and res.varieties:
                lines.append(f"{res.source_value} → {', '.join(res.varieties)}")
            elif res.warnings:
                lines.append(f"{res.source_value} → {res.status}: {'; '.join(res.warnings)}")
        self.resolved_selection_text.set("Selección resuelta: " + ("\n".join(lines) if lines else ""))
    def _filters(self):
        ctx=self.context_panel.context(); validate_context(ctx); d=self.remesa_panel.data(); period=Period(parse_user_date(d["desde"]),parse_user_date(d["hasta"])); validate_period(period)
        return DeliveryFilter(ctx, period, list(self.selected.get(0,"end")), d.get("socio"), d.get("categoria") or None)
    def _search(self):
        try:
            rows,summary,elapsed,total=self.deliveries.search(self._filters()); self.current_deliveries=rows; self.deliveries_panel.set_rows(rows); self.summary=summary; self.summary_panel.set_summary(summary)
            extra="" if total<=len(rows) else f" Se muestran las primeras {len(rows)}."
            self.status.set(f"{total} entregas cargadas en {elapsed:.2f} s.{extra}")
            if not total: messagebox.showinfo("Sin datos","No se han encontrado entregas para el periodo seleccionado.")
            self._refresh_action_states()
        except Exception as exc: messagebox.showerror("Error",str(exc))
    def _load_remesa(self):
        if self._has_pending_batch() and not self._confirm_discard_pending_batch(): return
        ctx=self.context_panel.context()
        if not (ctx.campana and ctx.empresa and ctx.cultivo):
            messagebox.showwarning("Contexto obligatorio", "Seleccione campaña, empresa y cultivo antes de cargar una remesa."); return
        try:
            items=self.remesas.list_remesas(ctx.campana, ctx.empresa, ctx.cultivo)
            selected=self._select_remesa_dialog(items, ctx)
            if not selected: return
            if isinstance(selected, list):
                if len(selected) > 1:
                    self._process_selected_remittances(selected)
                    return
                remesa_id=selected[0].remittance_id
            else:
                remesa_id=selected
            rem=self.remesas.get_remesa(remesa_id); self.current_remesa=rem; self.remesa_panel.load(rem.values)
            for k,v in rem.prices.items(): self.price_vars[k].set(str(v if v is not None else ""))
            self.apply_collection_var.set(parse_yes_no(rem.values.get("AplRec")))
            self.apply_transport_var.set(parse_yes_no(rem.values.get("AplTte")))
            self.apply_quality_var.set(parse_yes_no(rem.values.get("AplCal")))
            self.apply_globalgap_var.set(parse_yes_no(rem.values.get("AplGlobal")))
            self.apply_hectare_fee_var.set(parse_yes_no(rem.values.get("AplCHa")))
            self.apply_precalibrated_var.set(parse_yes_no(rem.values.get("AplPrecalibrado")))
            self._load_varieties(); self._restore_remesa_varieties(rem)
            self.deliveries_panel.clear(); self.summary_panel.clear(); self.current_deliveries=[]; self.current_calculation=None; self.current_group_benchmarks={}; self.calculation_valid=False
            self._refresh_action_states()
            if messagebox.askyesno("Buscar entregas", "¿Desea buscar las entregas correspondientes a esta remesa?"):
                self._search()
        except Exception as exc: messagebox.showerror("Error",str(exc))

    def _selected_remittance_from_values(self, values, ctx) -> SelectedRemittance:
        def parse_or_none(value):
            try:
                return parse_user_date(value) if value else None
            except Exception:
                return None
        return SelectedRemittance(
            remittance_id=int(values[0]),
            name=str(values[1] or ""),
            payment_date=parse_or_none(values[2]),
            period_from=parse_or_none(values[3]),
            period_to=parse_or_none(values[4]),
            category=str(values[5] or ""),
            liquidation_type=str(values[6] or ""),
            campaign=str(ctx.campana),
            company=str(ctx.empresa),
            crop=str(ctx.cultivo),
        )

    def _select_remesa_dialog(self, items, ctx):
        win=tk.Toplevel(self); win.title("Seleccionar remesa"); win.transient(self.winfo_toplevel()); win.grab_set(); win.geometry("980x500")
        ttk.Label(win,text=f"Campaña: {ctx.campana} | Empresa: {ctx.empresa} | Cultivo: {ctx.cultivo}").pack(anchor="w",padx=8,pady=4)
        query=tk.StringVar(); ttk.Entry(win,textvariable=query).pack(fill="x",padx=8,pady=4)
        selected_text=tk.StringVar(value="Remesas seleccionadas: 0")
        ttk.Label(win,textvariable=selected_text).pack(anchor="w",padx=8,pady=(0,4))
        cols=("IdREMESA","REMESA","FECHARE","PERIODO1","PERIODO2","CATEGORIA","TipoLiq")
        tree=ttk.Treeview(win,columns=cols,show="headings",selectmode="extended")
        [tree.heading(c,text=c) for c in cols]; [tree.column(c,width=130,anchor="w") for c in cols]; tree.pack(fill="both",expand=True,padx=8,pady=4)
        result={"items":None}
        def update_count(_=None):
            selected_text.set(f"Remesas seleccionadas: {len(tree.selection())}")
        def fill():
            tree.delete(*tree.get_children()); q=query.get().strip().upper()
            for row in items:
                hay=" ".join(str(row.get(k) or "") for k in ("IdREMESA","REMESA","CATEGORIA","TipoLiq")).upper()
                if not q or q in hay: tree.insert("","end",values=[row.get(c) or "" for c in cols])
            update_count()
        def select_all():
            tree.selection_set(tree.get_children()); update_count()
        def clear_selection():
            tree.selection_remove(tree.selection()); update_count()
        def load(_=None):
            sel=tree.selection()
            if not sel:
                messagebox.showwarning("Seleccionar remesa", "Seleccione al menos una remesa."); return
            remittances=[self._selected_remittance_from_values(tree.item(i,"values"), ctx) for i in sel]
            lines="\n".join(f"{r.remittance_id} - {r.name}" for r in remittances)
            if not messagebox.askyesno("Confirmar lote", f"Se van a procesar {len(remittances)} remesas:\n\n{lines}\n\nCada remesa se calculará de forma independiente.\nSe generará un único Excel resumen acumulado.\n\n¿Desea continuar?"):
                return
            result["items"]=remittances; win.destroy()
        query.trace_add("write", lambda *_: fill()); tree.bind("<Double-1>", load); tree.bind("<<TreeviewSelect>>", update_count); win.bind("<Return>", load); win.bind("<Escape>", lambda e: win.destroy())
        bf=ttk.Frame(win); bf.pack(fill="x",padx=8,pady=6)
        ttk.Button(bf,text="Seleccionar todas",command=select_all).pack(side="left",padx=4)
        ttk.Button(bf,text="Limpiar selección",command=clear_selection).pack(side="left",padx=4)
        ttk.Button(bf,text="Procesar seleccionadas",command=load).pack(side="right",padx=4)
        ttk.Button(bf,text="Cancelar",command=win.destroy).pack(side="right")
        fill(); win.wait_window(); return result["items"]

    def _batch_progress_dialog(self, total: int):
        win=tk.Toplevel(self); win.title("Procesando remesas"); win.transient(self.winfo_toplevel()); win.grab_set(); win.geometry("520x230"); win.resizable(False, False)
        general=tk.StringVar(value=f"Procesando remesa 0 de {total}")
        current=tk.StringVar(value="")
        phase=tk.StringVar(value="Fase: Preparando")
        members=tk.StringVar(value="Socios: -")
        bar=ttk.Progressbar(win, maximum=max(total, 1), mode="determinate")
        ttk.Label(win,textvariable=general,font=("TkDefaultFont",10,"bold")).pack(anchor="w",padx=12,pady=(12,4))
        ttk.Label(win,textvariable=current,wraplength=480).pack(anchor="w",padx=12,pady=4)
        ttk.Label(win,textvariable=phase).pack(anchor="w",padx=12,pady=4)
        ttk.Label(win,textvariable=members).pack(anchor="w",padx=12,pady=4)
        bar.pack(fill="x",padx=12,pady=8)
        ttk.Button(win,text="Cancelar después de la remesa actual",command=lambda: setattr(self,"batch_cancel_requested",True)).pack(anchor="e",padx=12,pady=8)
        def update(progress: BatchProgress):
            general.set(f"Procesando remesa {progress.current_index} de {progress.total_remittances}")
            current.set(progress.current_remittance_name or "")
            phase.set(f"Fase: {progress.message or progress.phase}")
            if progress.processed_members is not None and progress.total_members is not None:
                members.set(f"Socios: {progress.processed_members} de {progress.total_members}")
            bar["value"] = min(progress.current_index, progress.total_remittances)
            win.update_idletasks()
        return win, update

    def _process_selected_remittances(self, remittances: list[SelectedRemittance]) -> None:
        if self.batch_running:
            messagebox.showwarning("Lote en curso", "Ya hay un lote de remesas en ejecución."); return
        self.batch_running=True; self.batch_cancel_requested=False
        progress_win, update_progress = self._batch_progress_dialog(len(remittances))
        try:
            service=BatchRemittanceService(single_processor=self.process_single_remittance, exporter=export_batch_liquidation_summary, should_cancel=lambda: self.batch_cancel_requested)
            result=service.process(remittances, progress_callback=update_progress)
            self.current_batch_result=result
            self.current_batch_preview=self.persistence_service.prepare_batch_preview(result) if getattr(self,"persistence_enabled",False) else None
            self.current_batch_persisted=False; self.current_batch_save_result=None
            logger.info("[BatchPendingPersistence]\nexecution_id=%s\nremittances=%s\nvalid_remittances=%s\npending=true",getattr(self.current_batch_preview,"batch_execution_id",result.started_at.isoformat()),result.remittances_requested,result.remittances_completed)
            msg=(f"Las remesas se han calculado y quedan pendientes de guardar.\n\nRemesas procesadas: {result.remittances_requested}\nRemesas calculadas correctamente: {result.remittances_completed}\nRemesas con error de cálculo: {result.remittances_failed}\nBorradores generados: {result.drafts_generated}\nBorradores con error: {result.draft_errors}\n\nExcel acumulado:\n{result.aggregate_excel_path or 'No generado'}")
            messagebox.showinfo("Proceso terminado", msg)
            self.status.set(f"Lote terminado: {result.remittances_completed} correctas, {result.remittances_failed} con errores. Excel: {result.aggregate_excel_path or 'no generado'}")
            self._refresh_action_states()
        except PermissionError as exc:
            messagebox.showwarning("Archivo Excel abierto", str(exc))
        except Exception as exc:
            logger.exception("Error estructural procesando lote")
            messagebox.showerror("Procesando remesas", f"No se ha podido procesar el lote:\n{exc}")
        finally:
            self.batch_running=False
            if progress_win.winfo_exists(): progress_win.destroy()

    def process_single_remittance(self, remittance: SelectedRemittance, progress_callback=None, *, generate_individual_files: bool = True) -> SingleRemittanceBatchResult:
        def emit(phase, message):
            if progress_callback:
                progress_callback(BatchProgress(1, 1, remittance.remittance_id, remittance.name, phase, message=message))
        self.current_remesa=None; self.current_calculation=None; self.current_deliveries=[]; self.summary=None; self.current_group_benchmarks={}; self.calculation_valid=False
        emit("LOADING", "Cargando cabecera")
        rem=self.remesas.get_remesa(remittance.remittance_id); self.current_remesa=rem; self.remesa_panel.load(rem.values)
        for k,v in rem.prices.items(): self.price_vars[k].set(str(v if v is not None else ""))
        self.apply_collection_var.set(parse_yes_no(rem.values.get("AplRec"))); self.apply_transport_var.set(parse_yes_no(rem.values.get("AplTte"))); self.apply_quality_var.set(parse_yes_no(rem.values.get("AplCal"))); self.apply_globalgap_var.set(parse_yes_no(rem.values.get("AplGlobal"))); self.apply_hectare_fee_var.set(parse_yes_no(rem.values.get("AplCHa"))); self.apply_precalibrated_var.set(parse_yes_no(rem.values.get("AplPrecalibrado")))
        self._load_varieties(); self._restore_remesa_varieties(rem)
        emit("SEARCHING_DELIVERIES", "Buscando entregas")
        rows,summary,elapsed,total=self.deliveries.search(self._filters()); self.current_deliveries=list(rows); self.summary=summary
        emit("CALCULATING", "Calculando liquidación")
        calc_remesa=self._calculation_remesa(); calculation=self.calculations.calculate(list(rows), calc_remesa)
        self.current_calculation=calculation; self.calculation_valid=True
        if calculation and calculation.result:
            self.current_group_benchmarks=GroupBenchmarkService(GroupBenchmarkRepository(self.conn)).build_benchmarks(calculation.result.header, calculation.result.member_results)
            object.__setattr__(calculation.result, "variety_audit", tuple(calc_remesa.values.get("VARIEDAD_AUDIT", ())))
        generated=[]; self._batch_draft_errors=[]
        if generate_individual_files and calculation and calculation.result:
            emit("EXPORTING_FILES", "Generando archivos individuales")
            try:
                generated.append(export_liquidation_summary(calculation.result, self._output_dir()/"resumen_liquidaciones.xlsx"))
            except FileLockedError:
                pass
            generated.extend(export_hectare_fee_audit(calculation.result, self._output_dir()) or ())
            emit("GENERATING_PDFS", "Generando PDFs")
            members=self._premium_members()
            draft_paths=()
            if members:
                logger.info("[BatchDraftGenerationStarted] remittance_id=%s members=%s",remittance.remittance_id,len(members))
                draft_paths=self._export_premium_drafts(members,source="BATCH_DRAFT_EXPORT"); generated.extend(draft_paths)
        return SingleRemittanceBatchResult(remittance, calculation, calculation.member_count, calculation.delivery_count, self._output_dir(), tuple(generated),len(draft_paths) if generate_individual_files and calculation and calculation.result else 0,tuple(self._batch_draft_errors))

    def _restore_remesa_varieties(self, rem: Remesa) -> None:
        target=str(rem.values.get("VARIEDAD") or "").strip()
        self._clear_selected_varieties(invalidate=False)
        values=[v.strip() for v in target.split(",") if v.strip()]
        unresolved=[]
        for value in values:
            before=len(self.selected.get(0,"end")); self._add_source_variety(value, show_warning=False)
            if len(self.selected.get(0,"end")) == before:
                unresolved.append(value)
        if unresolved:
            messagebox.showwarning("Variedades", f"No se pudo resolver la variedad o grupo “{', '.join(unresolved)}”.")
    def _clear(self):
        self.current_remesa=None; self.remesa_panel.load({}); self.selected.delete(0,"end"); self.selected_source_items=[]; self.variety_resolutions=[]; self._refresh_resolved_selection_label(); self.deliveries_panel.clear(); self.summary_panel.clear(); self.summary=None; self.current_calculation=None; self.calculation_valid=False; self.current_deliveries=[]; self.status.set("Filtros/resultados limpiados"); self._refresh_action_states()

    def _new_remittance(self):
        if self._has_pending_batch() and not self._confirm_discard_pending_batch(): return
        self._clear()

    def _has_pending_batch(self):
        return self.current_batch_result is not None and not self.current_batch_persisted

    def _confirm_discard_pending_batch(self):
        answer=messagebox.askyesnocancel("Lote pendiente","Existe un lote calculado pendiente de guardar.\n\nSi continúa se perderá la selección calculada, aunque los archivos exportados permanecerán en disco.\n\nSí: Guardar ahora\nNo: Descartar\nCancelar: conservar el lote")
        if answer is None:return False
        if answer:return bool(self._save_batch_liquidations())
        self._discard_pending_batch(); return True

    def _discard_pending_batch(self):
        self.current_batch_result=None; self.current_batch_preview=None; self.current_batch_persisted=False; self.current_batch_save_result=None
        self.status.set("Lote pendiente descartado; los archivos exportados se conservan."); self._refresh_action_states()

    def close_application(self):
        if self._has_pending_batch() and not self._confirm_discard_pending_batch(): return
        self.preview_service.cleanup()
        self.winfo_toplevel().destroy()
    def _context_ready(self) -> bool:
        try:
            self._filters()
            return True
        except Exception:
            return False

    def _refresh_action_states(self) -> None:
        has_valid_context = self._context_ready()
        has_varieties = bool(self.selected.get(0,"end"))
        has_deliveries = bool(self.current_deliveries) or bool(self.deliveries_panel.visible_rows())
        calculation_ready = self.current_calculation is not None and self.calculation_valid
        premium_members_ready = bool(calculation_ready and self._premium_members())
        persistence_enabled = bool(getattr(self, "persistence_enabled", False))
        can_search = has_valid_context
        can_calculate = has_valid_context and has_varieties and has_deliveries
        can_persist_individual = persistence_enabled and calculation_ready and not self.current_calculation_persisted and self.current_batch_result is None
        can_persist_batch = persistence_enabled and self.current_batch_result is not None and not self.current_batch_persisted and self.current_batch_preview is not None and self.current_batch_preview.valid
        can_persist = can_persist_individual or can_persist_batch
        if hasattr(self, "action_buttons"):
            self.action_buttons["Buscar entregas"].configure(state="normal" if can_search else "disabled")
            self.action_buttons["Calcular liquidación"].configure(state="normal" if can_calculate else "disabled")
            self.action_buttons["Vista previa"].configure(state="normal" if calculation_ready else "disabled")
            self.action_buttons["Revisar lote"].configure(state="normal" if self.current_batch_result is not None else "disabled")
            self.action_buttons["Guardar liquidaciones"].configure(state="normal" if can_persist else "disabled")
            active_batch_ids = ()
            if persistence_enabled:
                active_batch_ids = tuple(
                    bid for bid in self.current_persisted_batch_ids
                    if (self.liquidation_repository.get_batch(bid) is not None
                        and self.liquidation_repository.get_batch(bid)["status"] == "ACTIVE")
                )
            can_void=persistence_enabled and bool(active_batch_ids or self.selected_history_batch_ids)
            self.action_buttons["Anular liquidación"].configure(state="normal" if can_void else "disabled")
            if "Exportar resumen de liquidación" in self.action_buttons: self.action_buttons["Exportar resumen de liquidación"].configure(state="normal" if calculation_ready else "disabled")
            if "Vista previa PDF" in self.action_buttons: self.action_buttons["Vista previa PDF"].configure(state="normal" if premium_members_ready else "disabled")
            if "Informe interno" in self.action_buttons: self.action_buttons["Informe interno"].configure(state="normal" if calculation_ready else "disabled")
        if hasattr(self,"persistence_status_text"):
            if self.current_batch_save_result and self.current_batch_save_result.failed: text=f"Lote guardado parcialmente: {self.current_batch_save_result.saved} de {self.current_batch_save_result.requested}."
            elif self.current_batch_persisted: text="Lote guardado correctamente."
            elif self.current_batch_preview: text=f"Lote pendiente de guardar: {sum(x.valid for x in self.current_batch_preview.remittances)} remesas."
            elif calculation_ready and not self.current_calculation_persisted: text="Liquidación individual pendiente de guardar."
            else: text="Sin cálculo pendiente."
            self.persistence_status_text.set(text)

    def _deliveries(self):
        return list(self.current_deliveries)

    def _calculate(self):
        try:
            deliveries = self._deliveries()
            if not deliveries:
                self._search(); deliveries = self._deliveries()
            if not deliveries:
                return
            remesa = self._calculation_remesa()
            self.current_calculation = self.calculations.calculate(deliveries, remesa)
            self.current_group_benchmarks = {}
            if self.current_calculation and self.current_calculation.result:
                self.current_group_benchmarks = GroupBenchmarkService(GroupBenchmarkRepository(self.conn)).build_benchmarks(self.current_calculation.result.header, self.current_calculation.result.member_results)
                object.__setattr__(self.current_calculation.result, "variety_audit", tuple(remesa.values.get("VARIEDAD_AUDIT", ())))
                audit_paths = export_hectare_fee_audit(self.current_calculation.result, self._output_dir())
                if audit_paths:
                    logger.info("Auditoría Cuota Ha generada: log=%s excel=%s", audit_paths[0], audit_paths[1])
            self.calculation_valid=True
            self.current_calculation_persisted=False
            self.summary_panel.set_calculation(self.current_calculation)
            self.status.set(f"Liquidación calculada: {self.current_calculation.member_count} socios, {format_decimal_es(self.current_calculation.net_kg, 3)} kg, importe comercial {format_currency_es(self.current_calculation.commercial_amount)}")
            self._refresh_action_states()
        except Exception as exc:
            messagebox.showerror("Error", str(exc))

    def _calculation_remesa(self):
        base = dict(self.current_remesa.values) if self.current_remesa else {}
        base.update({k:v.get() for k,v in self.price_vars.items()})
        ctx=self.context_panel.context(); data=self.remesa_panel.data()
        base.update({"CAMPAÑA":ctx.campana,"EMPRESA":ctx.empresa,"CULTIVO":ctx.cultivo,"REMESA":data.get("remesa"),"FECHARE":data.get("fecha_pago"),"PERIODO1":data.get("desde"),"PERIODO2":data.get("hasta"),"TipoLiq":data.get("tipo"),"CATEGORIA":data.get("categoria"),"IdSocio":data.get("socio"),"VARIEDAD":", ".join(self.selected_source_items or list(self.selected.get(0,"end"))),"AplRec":"S" if self.apply_collection_var.get() else "N","AplTte":"S" if self.apply_transport_var.get() else "N","AplCal":"S" if self.apply_quality_var.get() else "N","AplGlobal":"S" if self.apply_globalgap_var.get() else "N","AplCHa":"S" if self.apply_hectare_fee_var.get() else "N","AplPrecalibrado":"S" if self.apply_precalibrated_var.get() else "N"})
        base["VARIEDADES_RESUELTAS"] = list(self.selected.get(0,"end"))
        base["VARIEDAD_AUDIT"] = tuple(self.variety_resolutions)
        return Remesa(base)

    def _invalidate_master_changed(self):
        if getattr(self, "current_calculation", None):
            self.current_calculation = None
            logger.info("Invalidación de cálculo por cambio de maestro cuota Ha")
        self.calculation_valid = False
        self.status.set("La configuración de cuota Ha ha cambiado. Vuelva a calcular la liquidación.")
        self._refresh_action_states()

    def _invalidate_calculation(self):
        if getattr(self, "calculation_valid", False):
            self.calculation_valid=False; self.status.set("Los parámetros han cambiado. Vuelva a calcular la liquidación.")
        self._refresh_action_states()

    def _concept_text(self, value, status=None):
        if value is not None:
            return format_currency_es(value)
        if status and getattr(status, "value", "") == "not_applicable":
            return "No aplica"
        if status and getattr(status, "value", "") == "error":
            return "Error"
        return "Pendiente"

    def _sort_tree(self, tree, col, reverse=False):
        data = [(tree.set(k, col), k) for k in tree.get_children("")]
        def key(item):
            text = str(item[0]).replace(".", "").replace(",", ".").replace(" €", "").replace(" %", "")
            try: return float(text)
            except ValueError: return str(item[0]).lower()
        data.sort(key=key, reverse=reverse)
        for index, (_, k) in enumerate(data): tree.move(k, "", index)
        tree.heading(col, command=lambda: self._sort_tree(tree, col, not reverse))

    def _preview(self):
        if not self.deliveries_panel.visible_rows():
            return
        win=tk.Toplevel(self); win.title("Vista previa de liquidación"); win.transient(self.winfo_toplevel()); win.grab_set(); win.geometry("1280x760"); win.resizable(True, True)
        nb=ttk.Notebook(win); nb.pack(fill="both", expand=True, padx=8, pady=8)
        summary_tab=ttk.Frame(nb); members_tab=ttk.Frame(nb); detail_tab=ttk.Frame(nb)
        nb.add(summary_tab,text="Resumen"); nb.add(members_tab,text="Socios"); nb.add(detail_tab,text="Detalle")
        ctx=self.context_panel.context(); data=self.remesa_panel.data(); selected=', '.join(self.selected.get(0,'end')) or "Todas"; calc=self.current_calculation

        ident=ttk.LabelFrame(summary_tab,text="Identificación"); ident.grid(row=0,column=0,sticky="nsew",padx=6,pady=6)
        rows=[("Remesa",data.get('remesa')), ("Campaña",ctx.campana), ("Empresa",ctx.empresa), ("Cultivo",ctx.cultivo), ("Periodo",f"{format_display_date(data.get('desde'))} - {format_display_date(data.get('hasta'))}"), ("Fecha de pago",format_display_date(data.get('fecha_pago'))), ("Tipo de liquidación",data.get('tipo')), ("Categoría",data.get('categoria')), ("Socio",data.get('socio') or "Todos"), ("Variedades",selected), ("Precio €/ha activo", str(calc.result.hectare_fee_master.price_per_hectare).replace(".", ",") if calc and calc.result and calc.result.hectare_fee_master else ""), ("Cultivos sujetos a Cuota Ha", ", ".join(calc.result.hectare_fee_master.eligible_crops) if calc and calc.result and calc.result.hectare_fee_master else "")]
        for i,(k,v) in enumerate(rows):
            r=i//2; c=(i%2)*2; ttk.Label(ident,text=k,font=("TkDefaultFont",9,"bold")).grid(row=r,column=c,sticky="w",padx=4,pady=2); ttk.Label(ident,text=v,wraplength=360).grid(row=r,column=c+1,sticky="w",padx=4,pady=2)
        opts=ttk.LabelFrame(summary_tab,text="Opciones aplicadas"); opts.grid(row=1,column=0,sticky="ew",padx=6,pady=6)
        labels=[("Recolección",self.apply_collection_var), ("Transporte",self.apply_transport_var), ("Calidad",self.apply_quality_var), ("GlobalGAP",self.apply_globalgap_var), ("Cuota por hectárea",self.apply_hectare_fee_var), ("Precalibrado",self.apply_precalibrated_var)]
        for i,(name,var) in enumerate(labels): ttk.Label(opts,text=("✓ " if var.get() else "✗ ")+name).grid(row=i//3,column=i%3,sticky="w",padx=12,pady=3)
        origin=ttk.LabelFrame(summary_tab,text="Datos de origen"); origin.grid(row=0,column=1,rowspan=2,sticky="nsew",padx=6,pady=6)
        if self.summary:
            origin_rows=[("Entregas",self.summary.total_entregas),("Socios",self.summary.socios),("Variedades",self.summary.variedades),("Kilos netos efectivos",format_decimal_es(self.summary.kilos_netos,2)),("Primera fecha",self.summary.primera_fecha),("Última fecha",self.summary.ultima_fecha),("Entregas ya liquidadas",self.summary.liquidadas),("Registros con incidencias",self.summary.sin_variedad+self.summary.sin_socio_valido+self.summary.sin_categoria)]
            for i,(k,v) in enumerate(origin_rows): ttk.Label(origin,text=k,font=("TkDefaultFont",9,"bold")).grid(row=i,column=0,sticky="w",padx=4,pady=3); ttk.Label(origin,text=v).grid(row=i,column=1,sticky="w",padx=4,pady=3)
        econ=ttk.LabelFrame(summary_tab,text="Totales económicos"); econ.grid(row=2,column=0,columnspan=2,sticky="nsew",padx=6,pady=6)
        ecols=("Concepto","Importe"); etree=ttk.Treeview(econ,columns=ecols,show="headings",height=10); [etree.heading(c,text=c) for c in ecols]; etree.column("Concepto",width=220); etree.column("Importe",width=180,anchor="e"); etree.pack(fill="both",expand=True)
        totals=calc.result.totals if calc and calc.result else None
        concepts=[("Kilos netos efectivos", totals.net_kg if totals else None, None),("Importe comercial", calc.commercial_amount if calc else None, None),("Recolección", getattr(totals,'collection_amount',None), None),("Transporte", getattr(totals,'transport_amount',None), None),("Calidad", getattr(totals,'quality_amount',None), None),("GlobalGAP", getattr(totals,'globalgap_amount',None), None),("Cuota Ha", getattr(totals,'hectare_fee_amount',None), None),("Base imponible", getattr(totals,'taxable_base',None), None),("IVA", getattr(totals,'vat_amount',None), None),("Retención", getattr(totals,'withholding_amount',None), None),("Total", getattr(totals,'total_amount',None), None)]
        for name,val,st in concepts: etree.insert("","end",values=(name,self._concept_text(val,st)),tags=("strong",) if name in {"Base imponible","Total"} else ())
        etree.tag_configure("strong",font=("TkDefaultFont",9,"bold"))
        warn=ttk.LabelFrame(summary_tab,text="Advertencias"); warn.grid(row=3,column=0,columnspan=2,sticky="nsew",padx=6,pady=6)
        wlist=tk.Listbox(warn,height=4); wlist.pack(fill="both",expand=True)
        for msg in ((calc.warnings if calc else []) or ["Cálculo pendiente de ejecutar."]): wlist.insert("end",msg)
        summary_tab.columnconfigure(0,weight=1); summary_tab.columnconfigure(1,weight=1); summary_tab.rowconfigure(2,weight=1)

        mcols=("Nº socio","Socio","Variedad","Entregas","Neto efectivo","Neto comercial","Neto destrío","Neto podrido","Importe comercial","Recolección","Transporte","Calidad","GlobalGAP","Cuota Ha","Base imponible","IVA","Importe IVA","Importe después IVA","Retención","Importe Retención","Total","Precio medio final","Ha","Cuota anual","Kg totales Ha","€/kg Ha","Kg línea Ha","Cuota parcial Ha","Estado Ha")
        mtree=ttk.Treeview(members_tab,columns=mcols,show="headings");
        for c in mcols: mtree.heading(c,text=c,command=lambda c=c: self._sort_tree(mtree,c)); mtree.column(c,width=120,anchor="e" if c not in {"Socio","Variedad"} else "w")
        my=ttk.Scrollbar(members_tab,orient="vertical",command=mtree.yview); mx=ttk.Scrollbar(members_tab,orient="horizontal",command=mtree.xview); mtree.configure(yscrollcommand=my.set,xscrollcommand=mx.set); mtree.grid(row=0,column=0,sticky="nsew"); my.grid(row=0,column=1,sticky="ns"); mx.grid(row=1,column=0,sticky="ew"); members_tab.rowconfigure(0,weight=1); members_tab.columnconfigure(0,weight=1)
        dcols=("Nº socio","Socio","Variedad","Registro","Concepto","Coste_Recoleccion","SSocialRecoleccion","Manijeria","Recolección entrega","Coste_Trans","Kilos","Precio","Importe")
        dtree=ttk.Treeview(detail_tab,columns=dcols,show="headings");
        for c in dcols: dtree.heading(c,text=c,command=lambda c=c: self._sort_tree(dtree,c)); dtree.column(c,width=140,anchor="e" if c not in {"Socio","Variedad","Concepto"} else "w")
        dy=ttk.Scrollbar(detail_tab,orient="vertical",command=dtree.yview); dx=ttk.Scrollbar(detail_tab,orient="horizontal",command=dtree.xview); dtree.configure(yscrollcommand=dy.set,xscrollcommand=dx.set); dtree.grid(row=0,column=0,sticky="nsew"); dy.grid(row=0,column=1,sticky="ns"); dx.grid(row=1,column=0,sticky="ew"); detail_tab.rowconfigure(0,weight=1); detail_tab.columnconfigure(0,weight=1)
        if calc and calc.result:
            for m in calc.result.member_results:
                mtree.insert("","end",values=(m.member_id,m.member_name,m.variety,m.delivery_count,format_decimal_es(m.net_kg,2),format_decimal_es(m.commercial_kg,2),format_decimal_es(m.destruction_kg,2),format_decimal_es(m.rotten_kg,2),format_currency_es(m.commercial_amount),self._concept_text(m.collection_amount),self._concept_text(m.transport_amount),self._concept_text(m.quality_amount),self._concept_text(m.globalgap_amount),self._concept_text(m.hectare_fee_amount),self._concept_text(m.taxable_base),format_percentage_es(m.vat_rate or 0),self._concept_text(m.vat_amount),self._concept_text(m.amount_after_vat),format_percentage_es(m.withholding_rate or 0),self._concept_text(m.withholding_amount),self._concept_text(m.total_amount),format_price_es(m.final_average_price or 0),format_decimal_es(m.applicable_hectares,4),format_currency_es(m.hectare_fee_total_member),format_decimal_es(m.hectare_fee_total_effective_kg,2),format_price_es(m.hectare_fee_rate_per_kg or 0),format_decimal_es(m.net_kg,2),self._concept_text(m.hectare_fee_amount,m.hectare_fee_status),getattr(m.hectare_fee_status,"value",m.hectare_fee_status)))
                for g in m.grades:
                    if g.kilograms or g.price: dtree.insert("","end",values=(m.member_id,m.member_name,m.variety,"",g.label,"","","","","",format_decimal_es(g.kilograms,2),format_price_es(g.price),format_currency_es(g.amount)))
                for d in m.source_deliveries:
                    collection = d.collection_cost + d.social_security_collection + d.foreman_cost
                    dtree.insert("","end",values=(m.member_id,m.member_name,m.variety,d.registro,"Costes de entrada",format_currency_es(d.collection_cost),format_currency_es(d.social_security_collection),format_currency_es(d.foreman_cost),format_currency_es(collection),format_currency_es(d.transport_cost),format_decimal_es(d.effective_net_kg,2),"",""))
        buttons=ttk.Frame(win); buttons.pack(fill="x",padx=8,pady=(0,8))
        ttk.Button(buttons,text="Cerrar",command=win.destroy).pack(side="right",padx=3)
        ttk.Button(buttons,text="Copiar resumen",command=lambda: (win.clipboard_clear(), win.clipboard_append(f"{data.get('remesa')} - {format_currency_es(calc.commercial_amount) if calc else 'Pendiente'}"))).pack(side="right",padx=3)
        state="normal" if calc and calc.result else "disabled"
        ttk.Button(buttons,text="Vista previa Premium",command=self._export_premium_pdf,state=state).pack(side="right",padx=3)
        ttk.Button(buttons,text="Exportar resumen a Excel",command=self._export_liquidation_excel,state=state).pack(side="right",padx=3)
    def _output_dir(self):
        base=Path("C:/Liquidaciones/salidas/remesas") if Path("C:/").exists() else Path.cwd().parents[0]/"salidas"/"remesas"
        ctx=self.context_panel.context(); data=self.remesa_panel.data()
        p=base/safe_path_part(ctx.campana)/safe_path_part(ctx.cultivo)/(safe_path_part(data.get("remesa") or "remesa")); p.mkdir(parents=True,exist_ok=True); return p
    def _export_csv(self):
        path=self._output_dir()/"entregas_remesas.csv"
        with path.open("w",newline="",encoding="utf-8-sig") as f: w=csv.writer(f); w.writerow(COLUMNS); w.writerows(self.deliveries_panel.visible_rows())
        messagebox.showinfo("Exportación",f"CSV creado: {path}")
    def _export_excel(self):
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            messagebox.showerror("Exportación", "openpyxl no está instalado. Instale requirements.txt para exportar a Excel.")
            return
        wb=Workbook(); ws=wb.active; ws.title="Contexto"; ctx=self.context_panel.context(); data=self.remesa_panel.data()
        for row in [("campaña",ctx.campana),("empresa",ctx.empresa),("cultivo",ctx.cultivo),("periodo",f"{data['desde']} - {data['hasta']}"),("remesa",data['remesa'])]: ws.append(row)
        ws=wb.create_sheet("Entregas"); ws.append(COLUMNS); [ws.append(r) for r in self.deliveries_panel.visible_rows()]
        ws=wb.create_sheet("Resumen");
        if self.summary:
            for k,v in self.summary.__dict__.items(): ws.append((k,str(v)))
        path=self._output_dir()/"entregas_remesas.xlsx"; wb.save(path); messagebox.showinfo("Exportación",f"Excel creado: {path}")

    def _export_liquidation_excel(self):
        if not (self.current_calculation and self.current_calculation.result and self.calculation_valid): return
        try:
            path=export_liquidation_summary(self.current_calculation.result, self._output_dir()/"resumen_liquidaciones.xlsx")
        except FileLockedError as exc:
            logger.warning("Excel bloqueado: %s", exc.path)
            messagebox.showwarning(
                "Archivo Excel abierto",
                "No se puede actualizar el resumen porque el archivo está abierto o bloqueado.\n\n"
                f"Cierre el archivo:\n{exc.path}\n\n"
                "y vuelva a pulsar \"Exportar resumen de liquidación\"."
            )
            return
        except Exception as exc:
            logger.exception("Error exportando el resumen de liquidación")
            messagebox.showerror("Exportar resumen", f"No se ha podido generar el Excel:\n{exc}")
            return
        messagebox.showinfo("Exportación completada", f"El resumen se ha guardado en:\n{path}")

    def _premium_members(self):
        result = self.current_calculation.result if self.current_calculation and self.current_calculation.result else None
        if not result:
            return []
        # MemberLiquidation ya viene agrupado por socio y variedad desde el motor.
        # No deduplicamos por socio para no ocultar variedades dentro de la misma remesa.
        return sorted(result.member_results, key=lambda member: (member.member_id, member.variety or ""))

    def _premium_member_label(self, member) -> str:
        variety = f" · {member.variety}" if getattr(member, "variety", "") else ""
        return f"{member.member_id} - {member.member_name}{variety}"


    def _benchmark_for_member(self, member):
        result = self.current_calculation.result if self.current_calculation and self.current_calculation.result else None
        if not result:
            return None
        for key, benchmark in getattr(self, "current_group_benchmarks", {}).items():
            if key[0] == member.member_id and member.variety and member.variety.upper() in benchmark.varieties:
                return benchmark
        return None

    def _premium_member_path(self, member) -> Path:
        result = self.current_calculation.result
        vm = from_member_liquidation(result.header, member, group_benchmark=self._benchmark_for_member(member))
        return self._output_dir() / "borradores" / premium_member_filename(vm)

    def _write_premium_trace(self, *, mode: str, available, selected=None, paths=(), errors=0, error_text="") -> None:
        log_dir = Path.cwd() / "logs"
        log_dir.mkdir(parents=True, exist_ok=True)
        remesa = self.remesa_panel.data().get("remesa") or ""
        lines = [
            "===================================",
            "",
            "Generación PDF Premium",
            "",
            "Remesa:",
            str(remesa),
            "",
            "Socios disponibles:",
            "",
            *[str(member.member_id) for member in available],
            "",
            "Modo:",
            mode,
            "",
        ]
        if mode == "Todos":
            lines += ["Socios generados:", str(len(paths)), "", "Errores:", str(errors)]
        else:
            lines += ["Socio seleccionado:", str(selected.member_id if selected else ""), "", "Resultado:", "PDF generado correctamente" if paths and not errors else (error_text or "Error"), "", "Ruta:", str(paths[0]) if paths else ""]
        if error_text and mode == "Todos":
            lines += ["", "Detalle errores:", error_text]
        with (log_dir / "premium_pdf.log").open("a", encoding="utf-8") as fh:
            fh.write("\n".join(lines) + "\n")

    def _generate_premium_preview(self, member) -> Path:
        result = self.current_calculation.result
        vm = from_member_liquidation(result.header, member, group_benchmark=self._benchmark_for_member(member))
        path = self.preview_service.create_preview_path(
            member_id=member.member_id,
            member_name=member.member_name,
            remittance_name=result.header.remesa_name,
        )
        generate_liquidation_pdf(vm, path, document_mode=LiquidationDocumentMode.DRAFT)
        self.preview_service.open_preview(path)
        return path

    def _export_premium_draft(self, member, *, source="MANUAL_DRAFT_EXPORT") -> Path:
        result = self.current_calculation.result
        path=generate_liquidation_pdf(
            from_member_liquidation(result.header, member, group_benchmark=self._benchmark_for_member(member)),
            self._premium_member_path(member), document_mode=LiquidationDocumentMode.DRAFT,
        )
        if getattr(self,"persistence_enabled",False):
            from datetime import datetime, timezone
            self.liquidation_repository.record_exported_draft(remittance_id=int(result.header.remesa_id),
                recipient_member_id=int(member.member_id),member_name=member.member_name,
                campaign=str(result.header.campana),company=str(result.header.empresa),crop=str(result.header.cultivo),
                remittance_name=str(result.header.remesa_name),file_path=str(path),generated_at=datetime.now(timezone.utc).isoformat(),
                file_hash=hashlib.sha256(path.read_bytes()).hexdigest(),source=source)
        return path

    def _export_premium_drafts(self, members, *, source="MANUAL_DRAFT_EXPORT") -> tuple[Path, ...]:
        paths=[]
        for member in members:
            try:
                path=self._export_premium_draft(member,source=source); paths.append(path)
                logger.info("[BatchDraftGenerated] remittance_id=%s member_id=%s path=%s",self.current_calculation.result.header.remesa_id,member.member_id,path)
            except Exception as exc:
                logger.exception("[BatchDraftGenerationFailed] remittance_id=%s member_id=%s error=%s",self.current_calculation.result.header.remesa_id,member.member_id,exc)
                if source == "MANUAL_DRAFT_EXPORT": raise
                self._batch_draft_errors.append(f"ERROR DE BORRADOR: Remesa {self.current_calculation.result.header.remesa_id}: No se pudo generar el PDF borrador del socio {member.member_id}.")
        return tuple(paths)

    def _select_premium_members_dialog(self, members):
        selected = {"action": None, "member": members[0], "all": False}
        win=tk.Toplevel(self); win.title("Liquidaciones Premium"); win.transient(self.winfo_toplevel()); win.grab_set(); win.resizable(False, False)
        ttk.Label(win, text="Seleccione el socio").grid(row=0, column=0, columnspan=3, sticky="w", padx=12, pady=(12, 4))
        values=[self._premium_member_label(member) for member in members]
        combo_var=tk.StringVar(value=values[0])
        combo=ttk.Combobox(win, textvariable=combo_var, values=values, state="readonly", width=42)
        combo.grid(row=1, column=0, columnspan=3, sticky="ew", padx=12, pady=4)
        all_var=tk.BooleanVar(value=False)
        ttk.Checkbutton(win, text="Generar todos los socios", variable=all_var).grid(row=2, column=0, columnspan=3, sticky="w", padx=12, pady=4)
        ttk.Separator(win).grid(row=3, column=0, columnspan=3, sticky="ew", padx=12, pady=8)
        def finish(action):
            selected["action"] = action
            selected["all"] = all_var.get()
            selected["member"] = members[values.index(combo_var.get())]
            win.destroy()
        ttk.Button(win, text="Cancelar", command=lambda: finish("cancel")).grid(row=4, column=0, padx=8, pady=(0, 12))
        ttk.Button(win, text="Vista previa", command=lambda: finish("preview")).grid(row=4, column=1, padx=8, pady=(0, 12))
        ttk.Button(win, text="Generar PDF", command=lambda: finish("generate")).grid(row=4, column=2, padx=8, pady=(0, 12))
        win.wait_window()
        return selected

    def _export_premium_pdf(self):
        if not (self.current_calculation and self.current_calculation.result and self.calculation_valid): return
        members = self._premium_members()
        if not members:
            messagebox.showwarning("Liquidación Premium", "No hay socios disponibles en el cálculo actual.")
            return
        selection = {"action": "preview", "member": members[0], "all": False} if len(members) == 1 else self._select_premium_members_dialog(members)
        if selection["action"] == "cancel":
            return
        selected_members = members if selection["all"] and selection["action"] == "generate" else [selection["member"]]
        mode = "Todos" if selection["all"] and selection["action"] == "generate" else "Individual"
        try:
            if selection["action"] == "preview":
                paths = (self._generate_premium_preview(selection["member"]),)
            else:
                paths = self._export_premium_drafts(selected_members)
        except FileLockedError as exc:
            logger.warning("PDF Premium bloqueado: %s", exc.path)
            self._write_premium_trace(mode=mode, available=members, selected=selection["member"], errors=1, error_text=str(exc.path))
            messagebox.showwarning("PDF abierto", LOCKED_PDF_MESSAGE)
            return
        except Exception as exc:
            logger.exception("Error exportando Liquidación Premium")
            self._write_premium_trace(mode=mode, available=members, selected=selection["member"], errors=1, error_text=str(exc))
            messagebox.showerror("Liquidación Premium", f"No se ha podido generar el PDF Premium:\n{exc}")
            return
        self._write_premium_trace(mode=mode, available=members, selected=selection["member"], paths=paths)
        if selection["action"] == "preview":
            messagebox.showinfo("Vista previa PDF", "Se ha abierto una vista previa temporal. Este archivo no se ha guardado como documento definitivo.")
        else:
            messagebox.showinfo("Liquidación Premium", f"Se han generado {len(paths)} borradores PDF en:\n{self._output_dir()/'borradores'}")

    def _export_liquidation_pdf(self):
        if not (self.current_calculation and self.current_calculation.result and self.calculation_valid): return
        path=export_member_pdf(self.current_calculation.result, self._output_dir()/"liquidacion_socios.pdf")
        messagebox.showinfo("Exportación", f"PDF de liquidación creado: {path}")
